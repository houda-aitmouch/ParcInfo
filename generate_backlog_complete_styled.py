#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de Backlog Produit Complet Stylé - ParcInfo
Crée un backlog produit structuré avec style professionnel et toutes les fonctionnalités
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet

def get_complete_backlog_data():
    """Retourne toutes les données du backlog complet"""
    return [
        # === AUTHENTIFICATION ET UTILISATEURS ===
        {'Prio': 'P1', 'EPIC': 'Authentification', 'Feature': 'Interface de connexion', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux me connecter au système', 'RG': 'RG-AUTH-001: Champs identifiant et mot de passe obligatoires', 'US': 'US-AUTH-001: Saisir mon identifiant et mot de passe', 'Dépendance': '-', 'Statut': 'Terminé', 'Commentaire': 'Interface responsive et sécurisée', 'Feature Link': '/users/login/', 'US Link': '/users/login/'},
        {'Prio': 'P1', 'EPIC': 'Authentification', 'Feature': 'Gestion des sessions', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux rester connecté', 'RG': 'RG-AUTH-002: Session valide pendant 8h d\'inactivité', 'US': 'US-AUTH-002: Ma session reste active', 'Dépendance': 'Interface de connexion', 'Statut': 'Terminé', 'Commentaire': 'Sessions Django avec timeout', 'Feature Link': '/users/logout/', 'US Link': '/users/logout/'},
        {'Prio': 'P1', 'EPIC': 'Authentification', 'Feature': 'Redirection par rôle', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux être redirigé', 'RG': 'RG-AUTH-003: Redirection selon le groupe utilisateur', 'US': 'US-AUTH-003: Accéder à mon espace de travail', 'Dépendance': 'Interface de connexion', 'Statut': 'Terminé', 'Commentaire': 'Middleware de redirection', 'Feature Link': '/users/redirect-user/', 'US Link': '/users/redirect-user/'},
        {'Prio': 'P1', 'EPIC': 'Gestion Utilisateurs', 'Feature': 'Profil utilisateur', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux consulter mon profil', 'RG': 'RG-USER-001: Données personnelles modifiables', 'US': 'US-USER-001: Voir mes informations personnelles', 'Dépendance': 'Authentification', 'Statut': 'Terminé', 'Commentaire': 'Gestion des données personnelles', 'Feature Link': '/users/profil/', 'US Link': '/users/profil/'},
        {'Prio': 'P1', 'EPIC': 'Gestion Utilisateurs', 'Feature': 'Gestion des rôles', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux gérer les rôles', 'RG': 'RG-USER-002: 4 rôles distincts avec permissions', 'US': 'US-USER-002: Créer/modifier les groupes', 'Dépendance': 'Authentification', 'Statut': 'Terminé', 'Commentaire': 'Système de groupes Django', 'Feature Link': '/users/superadmin/', 'US Link': '/users/superadmin/'},
        
        # === DEMANDES D'ÉQUIPEMENT ===
        {'Prio': 'P1', 'EPIC': 'Demandes Équipement', 'Feature': 'Création de demande', 'Persona': 'Employé', 'Macro US': 'En tant qu\'employé, je veux créer une demande', 'RG': 'RG-DEM-001: Champs obligatoires selon le type', 'US': 'US-DEM-001: Remplir le formulaire de demande', 'Dépendance': 'Authentification', 'Statut': 'Terminé', 'Commentaire': 'Formulaire dynamique selon catégorie', 'Feature Link': '/demande_equipement/nouvelle/', 'US Link': '/demande_equipement/nouvelle/'},
        {'Prio': 'P1', 'EPIC': 'Demandes Équipement', 'Feature': 'Consultation des demandes', 'Persona': 'Employé', 'Macro US': 'En tant qu\'employé, je veux voir mes demandes', 'RG': 'RG-DEM-002: Seules mes demandes visibles', 'US': 'US-DEM-002: Lister mes demandes avec statuts', 'Dépendance': 'Création de demande', 'Statut': 'Terminé', 'Commentaire': 'Interface de consultation avec filtres', 'Feature Link': '/demande_equipement/', 'US Link': '/demande_equipement/'},
        {'Prio': 'P1', 'EPIC': 'Demandes Équipement', 'Feature': 'Approbation des demandes', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux approuver', 'RG': 'RG-DEM-003: Validation selon budget et disponibilité', 'US': 'US-DEM-003: Traiter les demandes en attente', 'Dépendance': 'Création de demande', 'Statut': 'Terminé', 'Commentaire': 'Workflow d\'approbation', 'Feature Link': '/demande_equipement/approuver/<id>/', 'US Link': '/demande_equipement/approuver/<id>/'},
        {'Prio': 'P1', 'EPIC': 'Demandes Équipement', 'Feature': 'Signature de décharge', 'Persona': 'Employé', 'Macro US': 'En tant qu\'employé, je veux signer ma décharge', 'RG': 'RG-DEM-004: Signature obligatoire avant réception', 'US': 'US-DEM-004: Signer électroniquement ma décharge', 'Dépendance': 'Approbation de demande', 'Statut': 'Terminé', 'Commentaire': 'Signature électronique sécurisée', 'Feature Link': '/demande_equipement/signer-decharge/<id>/', 'US Link': '/demande_equipement/signer-decharge/<id>/'},
        
        # === MATÉRIEL INFORMATIQUE ===
        {'Prio': 'P1', 'EPIC': 'Matériel Informatique', 'Feature': 'Gestion du parc', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux gérer le matériel', 'RG': 'RG-MAT-001: Inventaire complet avec codes uniques', 'US': 'US-MAT-001: Ajouter/modifier/supprimer du matériel', 'Dépendance': '-', 'Statut': 'Terminé', 'Commentaire': 'Codes d\'inventaire automatiques', 'Feature Link': '/materiel_informatique/', 'US Link': '/materiel_informatique/'},
        {'Prio': 'P1', 'EPIC': 'Matériel Informatique', 'Feature': 'Affectation d\'équipements', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux affecter', 'RG': 'RG-MAT-002: Un équipement par utilisateur maximum', 'US': 'US-MAT-002: Assigner un équipement à un utilisateur', 'Dépendance': 'Gestion du parc', 'Statut': 'Terminé', 'Commentaire': 'Traçabilité des affectations', 'Feature Link': '/materiel_informatique/affecter/', 'US Link': '/materiel_informatique/affecter/'},
        
        # === MATÉRIEL BUREAU ===
        {'Prio': 'P1', 'EPIC': 'Matériel Bureau', 'Feature': 'Gestion du mobilier', 'Persona': 'Gestionnaire Bureau', 'Macro US': 'En tant que gestionnaire bureau, je veux gérer', 'RG': 'RG-MAT-003: Inventaire bureau avec localisation', 'US': 'US-MAT-003: Gérer le mobilier de bureau', 'Dépendance': '-', 'Statut': 'Terminé', 'Commentaire': 'Gestion des espaces de travail', 'Feature Link': '/materiel_bureautique/', 'US Link': '/materiel_bureautique/'},
        {'Prio': 'P1', 'EPIC': 'Matériel Bureau', 'Feature': 'Affectation mobilier', 'Persona': 'Gestionnaire Bureau', 'Macro US': 'En tant que gestionnaire bureau, je veux affecter', 'RG': 'RG-MAT-004: Mobilier selon poste de travail', 'US': 'US-MAT-004: Assigner du mobilier aux employés', 'Dépendance': 'Gestion du mobilier', 'Statut': 'Terminé', 'Commentaire': 'Configuration poste de travail', 'Feature Link': '/materiel_bureautique/affecter/', 'US Link': '/materiel_bureautique/affecter/'},
        
        # === COMMANDES INFORMATIQUES ===
        {'Prio': 'P1', 'EPIC': 'Commandes Informatiques', 'Feature': 'Création de commande', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux créer des commandes', 'RG': 'RG-COM-001: Validation fournisseur et budget', 'US': 'US-COM-001: Créer une commande avec lignes', 'Dépendance': 'Gestion du parc', 'Statut': 'Terminé', 'Commentaire': 'Gestion des devis et budgets', 'Feature Link': '/commande_informatique/ajouter/', 'US Link': '/commande_informatique/ajouter/'},
        {'Prio': 'P1', 'EPIC': 'Commandes Informatiques', 'Feature': 'Suivi des commandes', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux suivre', 'RG': 'RG-COM-002: Statuts de commande obligatoires', 'US': 'US-COM-002: Voir l\'état de mes commandes', 'Dépendance': 'Création de commande', 'Statut': 'Terminé', 'Commentaire': 'Workflow de commande', 'Feature Link': '/commande_informatique/liste/', 'US Link': '/commande_informatique/liste/'},
        
        # === COMMANDES BUREAU ===
        {'Prio': 'P1', 'EPIC': 'Commandes Bureau', 'Feature': 'Création commande bureau', 'Persona': 'Gestionnaire Bureau', 'Macro US': 'En tant que gestionnaire bureau, je veux commander', 'RG': 'RG-COM-003: Validation fournisseur bureau', 'US': 'US-COM-003: Créer commande mobilier', 'Dépendance': 'Gestion du mobilier', 'Statut': 'Terminé', 'Commentaire': 'Catalogue fournisseurs bureau', 'Feature Link': '/commande_bureau/ajouter/', 'US Link': '/commande_bureau/ajouter/'},
        {'Prio': 'P1', 'EPIC': 'Commandes Bureau', 'Feature': 'Suivi commandes bureau', 'Persona': 'Gestionnaire Bureau', 'Macro US': 'En tant que gestionnaire bureau, je veux suivre', 'RG': 'RG-COM-004: Statuts spécifiques bureau', 'US': 'US-COM-004: Suivre commandes mobilier', 'Dépendance': 'Création commande bureau', 'Statut': 'Terminé', 'Commentaire': 'Gestion des délais livraison', 'Feature Link': '/commande_bureau/liste/', 'US Link': '/commande_bureau/liste/'},
        
        # === FOURNISSEURS ===
        {'Prio': 'P1', 'EPIC': 'Fournisseurs', 'Feature': 'Gestion catalogue', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux gérer les fournisseurs', 'RG': 'RG-FOU-001: Données fournisseur complètes', 'US': 'US-FOU-001: Ajouter/modifier fournisseurs', 'Dépendance': '-', 'Statut': 'Terminé', 'Commentaire': 'Base de données fournisseurs', 'Feature Link': '/fournisseurs/', 'US Link': '/fournisseurs/'},
        {'Prio': 'P1', 'EPIC': 'Fournisseurs', 'Feature': 'Catalogue produits', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux gérer les catalogues', 'RG': 'RG-FOU-002: Désignations et descriptions', 'US': 'US-FOU-002: Maintenir les catalogues produits', 'Dépendance': 'Gestion catalogue', 'Statut': 'Terminé', 'Commentaire': 'Classification produits', 'Feature Link': '/fournisseurs/catalogue/', 'US Link': '/fournisseurs/catalogue/'},
        
        # === LIVRAISONS ===
        {'Prio': 'P1', 'EPIC': 'Livraisons', 'Feature': 'Réception de livraisons', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux réceptionner', 'RG': 'RG-LIV-001: PV de réception obligatoire', 'US': 'US-LIV-001: Valider les livraisons reçues', 'Dépendance': 'Commandes', 'Statut': 'Terminé', 'Commentaire': 'Contrôle qualité réception', 'Feature Link': '/livraison/nouvelle/', 'US Link': '/livraison/nouvelle/'},
        {'Prio': 'P1', 'EPIC': 'Livraisons', 'Feature': 'Suivi des livraisons', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux suivre', 'RG': 'RG-LIV-002: Statuts de livraison', 'US': 'US-LIV-002: Voir l\'état des livraisons', 'Dépendance': 'Réception de livraisons', 'Statut': 'Terminé', 'Commentaire': 'Traçabilité complète', 'Feature Link': '/livraison/', 'US Link': '/livraison/'},
        
        # === NOTIFICATIONS ===
        {'Prio': 'P2', 'EPIC': 'Notifications', 'Feature': 'Système de notifications', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux recevoir des notifications', 'RG': 'RG-NOT-001: Notifications temps réel', 'US': 'US-NOT-001: Voir mes notifications', 'Dépendance': 'Demandes Équipement', 'Statut': 'Terminé', 'Commentaire': 'Notifications push et email', 'Feature Link': '/users/notifications-demandes/', 'US Link': '/users/notifications-demandes/'},
        {'Prio': 'P2', 'EPIC': 'Notifications', 'Feature': 'Notifications garantie', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux être alerté', 'RG': 'RG-NOT-002: Alertes 30 jours avant expiration', 'US': 'US-NOT-002: Recevoir alertes garantie', 'Dépendance': 'Matériel Informatique', 'Statut': 'En cours', 'Commentaire': 'Système d\'alertes automatiques', 'Feature Link': '/users/notifications-garantie/', 'US Link': '/users/notifications-garantie/'},
        
        # === DASHBOARDS ===
        {'Prio': 'P2', 'EPIC': 'Dashboard', 'Feature': 'Tableau de bord employé', 'Persona': 'Employé', 'Macro US': 'En tant qu\'employé, je veux voir mon dashboard', 'RG': 'RG-DASH-001: Données personnalisées par rôle', 'US': 'US-DASH-001: Accéder à mon tableau de bord', 'Dépendance': 'Authentification', 'Statut': 'Terminé', 'Commentaire': 'Vue d\'ensemble personnalisée', 'Feature Link': '/users/employe/', 'US Link': '/users/employe/'},
        {'Prio': 'P2', 'EPIC': 'Dashboard', 'Feature': 'Dashboard gestionnaire info', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux mon dashboard', 'RG': 'RG-DASH-002: Métriques informatiques', 'US': 'US-DASH-002: Voir les KPIs informatiques', 'Dépendance': 'Authentification', 'Statut': 'Terminé', 'Commentaire': 'Indicateurs de performance', 'Feature Link': '/users/gestionnaire_info/', 'US Link': '/users/gestionnaire_info/'},
        {'Prio': 'P2', 'EPIC': 'Dashboard', 'Feature': 'Dashboard gestionnaire bureau', 'Persona': 'Gestionnaire Bureau', 'Macro US': 'En tant que gestionnaire bureau, je veux mon dashboard', 'RG': 'RG-DASH-003: Métriques bureau', 'US': 'US-DASH-003: Voir les KPIs bureau', 'Dépendance': 'Authentification', 'Statut': 'Terminé', 'Commentaire': 'Indicateurs bureau', 'Feature Link': '/users/gestionnaire_bureau/', 'US Link': '/users/gestionnaire_bureau/'},
        {'Prio': 'P2', 'EPIC': 'Dashboard', 'Feature': 'Dashboard super admin', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux le dashboard complet', 'RG': 'RG-DASH-004: Vue globale du système', 'US': 'US-DASH-004: Accéder à toutes les données', 'Dépendance': 'Authentification', 'Statut': 'Terminé', 'Commentaire': 'Administration complète', 'Feature Link': '/users/superadmin/', 'US Link': '/users/superadmin/'},
        
        # === CHATBOT ===
        {'Prio': 'P2', 'EPIC': 'Chatbot', 'Feature': 'Assistant IA', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux poser des questions', 'RG': 'RG-CHAT-001: Réponses basées sur la base de connaissances', 'US': 'US-CHAT-001: Interroger l\'assistant IA', 'Dépendance': '-', 'Statut': 'Terminé', 'Commentaire': 'RAG avec base de connaissances', 'Feature Link': '/chatbot/', 'US Link': '/chatbot/'},
        {'Prio': 'P2', 'EPIC': 'Chatbot', 'Feature': 'Base de connaissances', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux maintenir la base', 'RG': 'RG-CHAT-002: Documentation à jour', 'US': 'US-CHAT-002: Mettre à jour la documentation', 'Dépendance': 'Assistant IA', 'Statut': 'Terminé', 'Commentaire': 'Système RAG avancé', 'Feature Link': '/chatbot/admin/', 'US Link': '/chatbot/admin/'},
        
        # === RAPPORTS ===
        {'Prio': 'P2', 'EPIC': 'Rapports', 'Feature': 'Export Excel', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux exporter des données', 'RG': 'RG-RAP-001: Formats Excel standardisés', 'US': 'US-RAP-001: Exporter les données en Excel', 'Dépendance': 'Gestion du parc', 'Statut': 'Terminé', 'Commentaire': 'Templates Excel personnalisés', 'Feature Link': '/materiel_informatique/export_excel/', 'US Link': '/materiel_informatique/export_excel/'},
        {'Prio': 'P2', 'EPIC': 'Rapports', 'Feature': 'Rapports de livraison', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux des rapports', 'RG': 'RG-RAP-002: Statistiques de livraison', 'US': 'US-RAP-002: Générer des rapports', 'Dépendance': 'Livraisons', 'Statut': 'En cours', 'Commentaire': 'Tableaux de bord analytiques', 'Feature Link': '/livraison/rapports/', 'US Link': '/livraison/rapports/'},
        
        # === RECHERCHE ===
        {'Prio': 'P2', 'EPIC': 'Recherche', 'Feature': 'Recherche globale', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux rechercher', 'RG': 'RG-RECH-001: Recherche multi-critères', 'US': 'US-RECH-001: Trouver rapidement des informations', 'Dépendance': '-', 'Statut': 'À faire', 'Commentaire': 'Recherche sémantique', 'Feature Link': '/users/search/', 'US Link': '/users/search/'},
        
        # === FONCTIONNALITÉS AVANCÉES ===
        {'Prio': 'P3', 'EPIC': 'Garanties', 'Feature': 'Suivi des garanties', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux suivre les garanties', 'RG': 'RG-GAR-001: Alertes automatiques', 'US': 'US-GAR-001: Gérer les garanties', 'Dépendance': 'Matériel Informatique', 'Statut': 'En cours', 'Commentaire': 'Système d\'alertes', 'Feature Link': '/users/dashboard-garantie/', 'US Link': '/users/dashboard-garantie/'},
        {'Prio': 'P3', 'EPIC': 'Archives', 'Feature': 'Archivage électronique', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux archiver', 'RG': 'RG-ARC-001: Conservation légale', 'US': 'US-ARC-001: Archiver les documents', 'Dépendance': 'Demandes Équipement', 'Statut': 'À faire', 'Commentaire': 'Archivage sécurisé', 'Feature Link': '/demande_equipement/archives/', 'US Link': '/demande_equipement/archives/'},
        {'Prio': 'P3', 'EPIC': 'API', 'Feature': 'API REST', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux une API', 'RG': 'RG-API-001: Documentation OpenAPI', 'US': 'US-API-001: Exposer les données via API', 'Dépendance': '-', 'Statut': 'À faire', 'Commentaire': 'API RESTful complète', 'Feature Link': '/api/', 'US Link': '/api/'},
        {'Prio': 'P3', 'EPIC': 'Sécurité', 'Feature': 'Audit des permissions', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux auditer', 'RG': 'RG-SEC-001: Traçabilité des actions', 'US': 'US-SEC-001: Voir les logs d\'audit', 'Dépendance': 'Authentification', 'Statut': 'À faire', 'Commentaire': 'Journalisation complète', 'Feature Link': '/audit/', 'US Link': '/audit/'},
        {'Prio': 'P3', 'EPIC': 'Mobile', 'Feature': 'Application mobile', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux accéder depuis mobile', 'RG': 'RG-MOB-001: Interface responsive', 'US': 'US-MOB-001: Consulter depuis mobile', 'Dépendance': 'API REST', 'Statut': 'À faire', 'Commentaire': 'Application mobile ou PWA', 'Feature Link': '/mobile/', 'US Link': '/mobile/'},
        {'Prio': 'P3', 'EPIC': 'Intégration', 'Feature': 'Intégration RH', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux intégrer avec les RH', 'RG': 'RG-INT-001: Synchronisation automatique', 'US': 'US-INT-001: Importer les utilisateurs depuis le système RH', 'Dépendance': 'API REST', 'Statut': 'À faire', 'Commentaire': 'Intégration LDAP/Active Directory', 'Feature Link': '/integration/rh/', 'US Link': '/integration/rh/'},
        {'Prio': 'P3', 'EPIC': 'BI', 'Feature': 'Business Intelligence', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux des analyses avancées', 'RG': 'RG-BI-001: Tableaux de bord interactifs', 'US': 'US-BI-001: Analyser les tendances', 'Dépendance': 'Rapports', 'Statut': 'À faire', 'Commentaire': 'Intégration Power BI ou Tableau', 'Feature Link': '/bi/', 'US Link': '/bi/'}
    ]

def apply_styles_to_worksheet(ws, title="Backlog Produit"):
    """Applique des styles professionnels à la feuille de calcul"""
    
    # Styles de base
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre principal
    ws['A1'] = f"📋 {title} - ParcInfo"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
    ws.merge_cells('A1:L1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Sous-titre avec date
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws.merge_cells('A2:L2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # En-têtes de colonnes
    headers = ['Prio', 'EPIC', 'Feature', 'Persona', 'Macro US', 'RG', 'US', 'Dépendance', 'Statut', 'Commentaire', 'Feature Link', 'US Link']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 8,   # Prio
        'B': 20,  # EPIC
        'C': 25,  # Feature
        'D': 15,  # Persona
        'E': 40,  # Macro US
        'F': 20,  # RG
        'G': 25,  # US
        'H': 15,  # Dépendance
        'I': 12,  # Statut
        'J': 30,  # Commentaire
        'K': 20,  # Feature Link
        'L': 20   # US Link
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Hauteur de ligne pour les en-têtes
    ws.row_dimensions[4].height = 30
    
    return ws

def apply_priority_colors(ws, start_row=5):
    """Applique des couleurs selon la priorité et le statut"""
    
    # Couleurs par priorité
    priority_colors = {
        'P1': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),  # Rouge clair
        'P2': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # Jaune clair
        'P3': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')   # Bleu clair
    }
    
    # Couleurs par statut
    status_colors = {
        'Terminé': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),  # Vert
        'En cours': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),  # Jaune
        'À faire': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')    # Rouge
    }
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Couleur de priorité pour la première colonne
            if col == 1 and cell.value in priority_colors:
                cell.fill = priority_colors[cell.value]
                cell.font = Font(bold=True)
            
            # Couleur de statut pour la colonne statut
            if col == 9 and cell.value in status_colors:
                cell.fill = status_colors[cell.value]
                cell.font = Font(bold=True)

def generate_complete_styled_backlog():
    """Génère le backlog produit complet avec style professionnel"""
    
    # Récupération des données complètes
    backlog_data = get_complete_backlog_data()
    df = pd.DataFrame(backlog_data)
    
    # Tri par priorité et EPIC
    priority_order = {'P1': 1, 'P2': 2, 'P3': 3}
    df['Prio_Order'] = df['Prio'].map(priority_order)
    df = df.sort_values(['Prio_Order', 'EPIC', 'Feature'])
    df = df.drop('Prio_Order', axis=1)
    
    # Création du workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog Produit"
    
    # Application des styles
    apply_styles_to_worksheet(ws)
    
    # Ajout des données
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    
    # Application des couleurs
    apply_priority_colors(ws, start_row=5)
    
    # Création des autres feuilles (simplifiées pour cet exemple)
    create_personas_sheet(wb)
    create_epics_sheet(wb)
    
    # Suppression de la feuille par défaut si elle existe
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sauvegarde
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Backlog_Produit_ParcInfo_Complet_Styled_{timestamp}.xlsx"
    wb.save(filename)
    
    print(f"✅ Backlog produit complet stylé généré : {filename}")
    print(f"📊 {len(df)} fonctionnalités avec style professionnel")
    print(f"🎨 Couleurs par priorité et statut appliquées")
    print(f"📋 {len(df['EPIC'].unique())} EPICs organisés")
    print(f"👥 {len(df['Persona'].unique())} personas identifiés")
    
    return filename

# Fonctions simplifiées pour les autres feuilles
def create_personas_sheet(wb):
    """Crée la feuille des personas avec style"""
    ws = wb.create_sheet("Personas")
    ws['A1'] = "👥 Personas - ParcInfo"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')

def create_epics_sheet(wb):
    """Crée la feuille des EPICs avec style"""
    ws = wb.create_sheet("EPICs")
    ws['A1'] = "📋 EPICs - ParcInfo"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center')

if __name__ == "__main__":
    generate_complete_styled_backlog()
