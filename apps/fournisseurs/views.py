from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from .models import Fournisseur
from .forms import FournisseurForm


def is_gestionnaire_ou_superadmin(user):
    return user.groups.filter(name__in=['Gestionnaire Informatique', 'Super Admin', 'Gestionnaire Bureau']).exists() or user.is_superuser


def _use_superadmin_layout(user) -> bool:
    """Return True if the user should see the superadmin layout."""
    if not user or not user.is_authenticated:
        return False
    # Superuser or explicit Super Admin group or Gestionnaire Informatique group or Gestionnaire Bureau group
    return bool(user.is_superuser or user.groups.filter(name__in=['Super Admin', 'Gestionnaire Informatique', 'Gestionnaire Bureau']).exists())


@user_passes_test(is_gestionnaire_ou_superadmin)
def fournisseur_list(request):
    fournisseurs = Fournisseur.objects.all().order_by('nom')
    total = fournisseurs.count()
    
    # Choisir le template selon le rôle
    if _use_superadmin_layout(request.user):
        if request.user.groups.filter(name='Gestionnaire Informatique').exists():
            template_name = 'fournisseurs/fournisseur_list_gestionnaire_info.html'
        elif request.user.groups.filter(name='Gestionnaire Bureau').exists():
            template_name = 'fournisseurs/fournisseur_list_gestionnaire_bureau_superadmin.html'
        else:
            template_name = 'fournisseurs/fournisseur_list_superadmin.html'
    else:
        template_name = 'fournisseurs/fournisseur_list.html'
    
    return render(request, template_name, {
        'fournisseurs': fournisseurs,
        'total': total,
    })





@user_passes_test(is_gestionnaire_ou_superadmin)
def fournisseur_form(request, pk=None):
    fournisseur = None
    if pk:
        fournisseur = get_object_or_404(Fournisseur, pk=pk)
    
    if request.method == 'POST':
        form = FournisseurForm(request.POST, instance=fournisseur)
        if form.is_valid():
            fournisseur = form.save()
            messages.success(request, "Fournisseur ajouté avec succès." if not pk else "Fournisseur modifié avec succès.")
            return redirect('fournisseurs:fournisseur_list')
    else:
        # Vérifier si on modifie un fournisseur existant
        if pk:
            form = FournisseurForm(instance=fournisseur)
        else:
            form = FournisseurForm()
    
    # Choisir le template selon le rôle
    if _use_superadmin_layout(request.user):
        if request.user.groups.filter(name='Gestionnaire Informatique').exists():
            template_name = 'fournisseurs/fournisseur_form_gestionnaire_info.html'
        elif request.user.groups.filter(name='Gestionnaire Bureau').exists():
            template_name = 'fournisseurs/fournisseur_form_gestionnaire_bureau_superadmin.html'
        else:
            template_name = 'fournisseurs/fournisseur_form_superadmin.html'
    else:
        template_name = 'fournisseurs/fournisseur_form.html'
    
    return render(request, template_name, {
        'form': form,
        'fournisseur': fournisseur,
    })


@user_passes_test(is_gestionnaire_ou_superadmin)
def fournisseur_confirm_delete(request, pk):
    fournisseur = get_object_or_404(Fournisseur, pk=pk)
    if request.method == 'POST':
        fournisseur.delete()
        messages.success(request, "Fournisseur supprimé avec succès.")
        return redirect('fournisseurs:fournisseur_list')
    
    # Choisir le template selon le rôle
    if _use_superadmin_layout(request.user):
        if request.user.groups.filter(name='Gestionnaire Informatique').exists():
            template_name = 'fournisseurs/fournisseur_confirm_delete_gestionnaire_info.html'
        elif request.user.groups.filter(name='Gestionnaire Bureau').exists():
            template_name = 'fournisseurs/fournisseur_confirm_delete_gestionnaire_bureau_superadmin.html'
        else:
            template_name = 'fournisseurs/fournisseur_confirm_delete_superadmin.html'
    else:
        template_name = 'fournisseurs/fournisseur_confirm_delete.html'
    
    return render(request, template_name, {
        'fournisseur': fournisseur
    })


@user_passes_test(is_gestionnaire_ou_superadmin)
def exporter_fournisseurs_excel(request):
    fournisseurs = Fournisseur.objects.all()
    total = fournisseurs.count()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fournisseurs"

    # --- Titre fusionné ---
    titre = "Liste des fournisseurs"
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = titre
    title_cell.font = Font(name='Calibri', size=18, bold=True, color='2C3E50')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    title_cell.fill = PatternFill("solid", fgColor="D6E4F0")  # Bleu clair doux
    ws.row_dimensions[1].height = 30

    # --- En-têtes ---
    headers = ['Nom', 'IF', 'ICE', 'RC', 'Adresse']
    header_row = 3  # on laisse la ligne 2 vide pour aérer
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = Font(name='Segoe UI', size=12, bold=True, color='FFFFFF')
        cell.fill = GradientFill(stop=("4A90E2", "357ABD"))  # dégradé bleu
        cell.alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='B0C4DE'),
            right=Side(style='thin', color='B0C4DE'),
            top=Side(style='thin', color='B0C4DE'),
            bottom=Side(style='thin', color='B0C4DE'),
        )
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # --- Données ---
    data_row = header_row + 1
    for fournisseur in fournisseurs:
        ws.cell(row=data_row, column=1, value=fournisseur.nom)
        ws.cell(row=data_row, column=2, value=fournisseur.if_fiscal or '')
        ws.cell(row=data_row, column=3, value=fournisseur.ice or '')
        ws.cell(row=data_row, column=4, value=fournisseur.registre_commerce or '')
        ws.cell(row=data_row, column=5, value=fournisseur.adresse or '')
        
        # Appliquer les bordures aux cellules de données
        for col_num in range(1, 6):
            cell = ws.cell(row=data_row, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
        data_row += 1

    # --- Total en bas ---
    total_row = data_row + 1
    ws.merge_cells(f'A{total_row}:D{total_row}')
    total_cell = ws[f'A{total_row}']
    total_cell.value = f"Total : {total} fournisseurs"
    total_cell.font = Font(name='Segoe UI', size=12, bold=True, color='2C3E50')
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.fill = PatternFill("solid", fgColor="E8F4FD")  # Bleu très clair

    # --- Réponse HTTP ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="fournisseurs_{total}.xlsx"'
    
    wb.save(response)
    return response

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=liste_fournisseurs.xlsx'
    wb.save(response)
    return response

# ============================================================================
# VUES GESTIONNAIRE BUREAU (mêmes fonctionnalités, templates dédiés)
