from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from django.http import HttpResponse
from django import forms as django_forms  # Import django.forms avec un alias

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .models import Equipement, Materiel

# --- Formulaires ---

class MaterielForm(django_forms.ModelForm):
    class Meta:
        model = Materiel
        fields = '__all__'  # Tu peux spécifier une liste de champs si tu veux

# Si tu as un EquipementForm dans forms.py, importe-le
from .forms import EquipementForm  # Assure-toi que ce formulaire existe dans forms.py


# --- Vérifications d'accès ---

def is_gestionnaire_ou_superadmin(user):
    return user.is_authenticated and (
        user.groups.filter(name='Super Admin').exists() or
        user.groups.filter(name='Gestionnaire Informatique').exists()
    )

def is_superadmin_or_gestionnaire_bureau(user):
    return user.is_authenticated and (
        user.is_superuser or
        user.groups.filter(name='Gestionnaire Bureau').exists()
    )


# --- Vues Equipement ---

@user_passes_test(is_gestionnaire_ou_superadmin)
def equipement_list(request):
    equipements = Equipement.objects.all()
    return render(request, 'equipements/list.html', {'equipements': equipements})

@user_passes_test(is_gestionnaire_ou_superadmin)
def equipement_create(request):
    if request.method == 'POST':
        form = EquipementForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('equipements:equipement_list')
    else:
        form = EquipementForm()
    return render(request, 'equipements/equipement_form.html', {'form': form})

@user_passes_test(is_gestionnaire_ou_superadmin)
def equipement_update(request, pk):
    equipement = get_object_or_404(Equipement, pk=pk)
    if request.method == 'POST':
        form = EquipementForm(request.POST, instance=equipement)
        if form.is_valid():
            form.save()
            return redirect('equipements:equipement_list')
    else:
        form = EquipementForm(instance=equipement)
    return render(request, 'equipements/equipement_form.html', {'form': form})

@user_passes_test(is_gestionnaire_ou_superadmin)
def equipement_delete(request, pk):
    equipement = get_object_or_404(Equipement, pk=pk)
    if request.method == 'POST':
        equipement.delete()
        messages.success(request, "L'équipement a été supprimé avec succès.")
        return redirect('equipements:equipement_list')
    return render(request, 'equipements/equipement_confirm_delete.html', {'equipement': equipement})

@user_passes_test(is_gestionnaire_ou_superadmin)
def export_excel_equipement(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Liste des équipements"

    headers = [
        "N° Série", "Code N° Inventaire", "Date de service", "Désignation", "Description",
        "Prix HT (MAD)", "Fournisseur", "N° Facture", "Garantie", "Statut",
        "Public", "Lieu", "Utilisateur", "Observations"
    ]

    header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Entêtes
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = border

    # Données
    for row_num, e in enumerate(Equipement.objects.all(), start=2):
        values = [
            e.sn_article,
            e.code_inventaire,
            e.date_service.strftime("%d/%m/%Y") if e.date_service else "",
            e.designation,
            e.description,
            float(e.prix_ht_mad) if e.prix_ht_mad else "",
            e.fournisseur,
            e.numero_facture,
            e.date_garantie.strftime("%d/%m/%Y") if e.date_garantie else "",
            e.get_statut_display(),
            "Oui" if e.is_public else "Non",
            e.get_lieu_affectation_display() if e.lieu_affectation else "",
            str(e.utilisateur) if e.utilisateur else "",
            e.observations
        ]
        for col_num, value in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border

            # Coloration conditionnelle sur le statut
            if headers[col_num - 1] == "Statut":
                val = value.lower() if isinstance(value, str) else ""
                if val == "en panne":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair
                elif val == "maintenance":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Jaune
                elif val == "affecté":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair

    # Ajustement largeur colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="equipements_stylés.xlsx"'
    wb.save(response)
    return response


# --- Vues Matériel Bureau ---
@user_passes_test(is_superadmin_or_gestionnaire_bureau)
def materiel_list(request):
    materiels = Materiel.objects.all()
    return render(request, 'materiel/materiel_list.html', {'materiels': materiels})

@user_passes_test(is_superadmin_or_gestionnaire_bureau)
def materiel_create(request):
    if request.method == 'POST':
        form = MaterielForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Matériel ajouté avec succès.")
            return redirect('materiel:materiel_list')
    else:
        form = MaterielForm()
    return render(request, 'materiel/materiel_form.html', {'form': form})
@user_passes_test(is_superadmin_or_gestionnaire_bureau)
def materiel_update(request, materiel_id):
    materiel = get_object_or_404(Materiel, id=materiel_id)
    if request.method == 'POST':
        form = MaterielForm(request.POST, instance=materiel)
        if form.is_valid():
            form.save()
            messages.success(request, "Matériel modifié avec succès.")
            return redirect('materiel:materiel_list')
    else:
        form = MaterielForm(instance=materiel)
    return render(request, 'materiel/materiel_form.html', {'form': form, 'materiel': materiel})
@user_passes_test(is_superadmin_or_gestionnaire_bureau)
def materiel_delete(request, materiel_id):
    materiel = get_object_or_404(Materiel, id=materiel_id)
    if request.method == 'POST':
        materiel.delete()
        messages.success(request, "Matériel supprimé avec succès.")
        return redirect('materiel:materiel_list')
    return render(request, 'materiel/materiel_confirm_delete.html', {'materiel': materiel})
@user_passes_test(is_superadmin_or_gestionnaire_bureau)
def export_excel_materiel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Matériel de Bureau"

    headers = [
        "Code Inventaire", "Date Mise en Service", "Description", "Désignation",
        "Prix HT en MAD", "Fournisseur", "État Article", "Lieu d'Affectation", "Observations"
    ]

    header_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # En-têtes
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = border

    # Données
    for row_num, m in enumerate(Materiel.objects.all(), start=2):
        row = [
            m.code_inventaire,
            m.date_service.strftime("%d/%m/%Y") if m.date_service else "",
            m.description,
            m.designation,
            float(m.prix_ht_mad) if m.prix_ht_mad else "",
            m.fournisseur,
            m.etat_article.capitalize() if m.etat_article else "",
            m.lieu_affectation,
            m.observations
        ]
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = border

            # Coloration conditionnelle sur l'état
            if headers[col_num - 1] == "État Article":
                val = value.lower() if isinstance(value, str) else ""
                if val == "opérationnel":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Vert clair
                elif val == "réparation":
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Jaune
                elif val == "réforme":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Rouge clair

    # Ajuster largeur colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 5

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="materiel_bureau.xlsx"'
    wb.save(response)
    return response