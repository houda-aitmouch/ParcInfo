# apps/commande/views.py
from django.forms import modelformset_factory
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.contrib import messages
from django.core.serializers import serialize
import json
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.http import HttpResponse
from django.core.exceptions import PermissionDenied

from .models import CommandeBureau, DesignationBureau, DescriptionBureau, LigneCommandeBureau
from apps.fournisseurs.models import Fournisseur
from .forms import CommandeBureauForm, LigneCommandeBureauFormSet, LigneCommandeBureauForm

logger = logging.getLogger(__name__)

def is_gestionnaire_ou_superadmin(user):
    return user.groups.filter(name__in=['Gestionnaire Bureau', 'Super Admin']).exists()

def is_gestionnaire_bureau(user):
    return user.groups.filter(name='Gestionnaire Bureau').exists()

def is_superadmin(user):
    return user.is_superuser

def liste_commandes(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    commandes = CommandeBureau.objects.prefetch_related('lignes__designation', 'lignes__description')
    context = {
        'commandes': commandes
    }
    return render(request, 'commande_bureau/liste_commandes.html', context)

def liste_commandes_superadmin(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    commandes = CommandeBureau.objects.prefetch_related('lignes__designation', 'lignes__description')
    context = { 'commandes': commandes }
    return render(request, 'commande_bureau/liste_commandes_superadmin.html', context)

def liste_commandes_gestionnaire_info(request):
    if not (is_gestionnaire_ou_superadmin(request.user) or is_superadmin(request.user)):
        raise PermissionDenied
    commandes = CommandeBureau.objects.prefetch_related('lignes__designation', 'lignes__description')
    context = { 'commandes': commandes }
    return render(request, 'commande_bureau/liste_commandes_gestionnaire_info.html', context)

def liste_commandes_gestionnaire_bureau(request):
    if not (is_gestionnaire_bureau(request.user) or is_superadmin(request.user)):
        raise PermissionDenied
    commandes = CommandeBureau.objects.prefetch_related('lignes__designation', 'lignes__description')
    context = { 'commandes': commandes }
    return render(request, 'commande_bureau/liste_commandes_gestionnaire_bureau.html', context)

def ajouter_commande_gestionnaire_bureau(request):
    if not (is_gestionnaire_bureau(request.user) or is_superadmin(request.user)):
        raise PermissionDenied
    if request.method == 'POST':
        try:
            # Traitement des données du formulaire principal
            commande_data = {
                'mode_passation': request.POST.get('mode_passation'),
                'numero_commande': request.POST.get('numero_commande'),
                'fournisseur': request.POST.get('fournisseur'),
                'date_commande': request.POST.get('date_commande'),
                'date_reception': request.POST.get('date_reception'),
                'numero_facture': request.POST.get('numero_facture'),
                'duree_garantie_valeur': request.POST.get('duree_garantie_valeur'),
                'duree_garantie_unite': request.POST.get('duree_garantie_unite'),
            }

            commande_form = CommandeBureauForm(commande_data)

            if commande_form.is_valid():
                commande = commande_form.save()

                # Traitement des lignes de commande
                lignes_data = []
                i = 0
                while f'lignes-{i}-designation' in request.POST:
                    ligne_data = {
                        'designation_id': request.POST.get(f'lignes-{i}-designation'),
                        'description_id': request.POST.get(f'lignes-{i}-description'),
                        'quantite': request.POST.get(f'lignes-{i}-quantite'),
                        'prix_unitaire': request.POST.get(f'lignes-{i}-prix_unitaire'),
                    }

                    # Validation des données de ligne
                    if all([ligne_data['designation_id'], ligne_data['description_id'],
                            ligne_data['quantite'], ligne_data['prix_unitaire']]):
                        try:
                            LigneCommandeBureau.objects.create(
                                commande=commande,
                                designation_id=ligne_data['designation_id'],
                                description_id=ligne_data['description_id'],
                                quantite=int(ligne_data['quantite']),
                                prix_unitaire=float(ligne_data['prix_unitaire'])
                            )
                        except (ValueError, TypeError) as e:
                            logger.error(f"Erreur lors de la création de la ligne {i}: {e}")
                            messages.error(request, f"Erreur dans la ligne {i + 1}: données invalides")

                    i += 1

                if i > 0:  # Au moins une ligne créée
                    messages.success(request, f'Commande {commande.numero_commande} créée avec succès!')
                    return redirect('commandes_bureau:liste_commandes_gestionnaire_bureau')
                else:
                    commande.delete()  # Supprimer la commande si aucune ligne n'est valide
                    messages.error(request, 'Aucune ligne de commande valide trouvée')
            else:
                messages.error(request, 'Erreur dans les données de la commande')

        except Exception as e:
            logger.error(f"Erreur lors de la création de la commande: {e}")
            messages.error(request, 'Erreur lors de la création de la commande')

    # GET request ou erreur dans POST
    designations = DesignationBureau.objects.all().order_by('nom')
    fournisseurs = Fournisseur.objects.all().order_by('nom')

    # Sérialiser les désignations pour JavaScript
    designations_json = json.dumps([
        {'id': d.id, 'nom': d.nom} for d in designations
    ])

    return render(request, 'commande_bureau/commande_form_gestionnaire_bureau.html', {
        'form': CommandeBureauForm(),
        'designations': designations,
        'designations_json': designations_json,
        'fournisseurs': fournisseurs,
    })

def modifier_commande_gestionnaire_bureau(request, pk):
    if not is_gestionnaire_bureau(request.user):
        raise PermissionDenied
    commande = get_object_or_404(CommandeBureau, pk=pk)
    
    if request.method == 'POST':
        try:
            commande_data = {
                'mode_passation': request.POST.get('mode_passation'),
                'numero_commande': request.POST.get('numero_commande'),
                'fournisseur': request.POST.get('fournisseur'),
                'date_commande': request.POST.get('date_commande'),
                'date_reception': request.POST.get('date_reception'),
                'numero_facture': request.POST.get('numero_facture'),
                'duree_garantie_valeur': request.POST.get('duree_garantie_valeur'),
                'duree_garantie_unite': request.POST.get('duree_garantie_unite'),
            }

            commande_form = CommandeBureauForm(commande_data, instance=commande)

            if commande_form.is_valid():
                commande = commande_form.save()
                messages.success(request, f'Commande {commande.numero_commande} modifiée avec succès!')
                return redirect('commandes_bureau:liste_commandes_gestionnaire_bureau')
            else:
                messages.error(request, 'Erreur dans les données de la commande')

        except Exception as e:
            logger.error(f"Erreur lors de la modification de la commande: {e}")
            messages.error(request, 'Erreur lors de la modification de la commande')

    # GET request ou erreur dans POST
    designations = DesignationBureau.objects.all().order_by('nom')
    fournisseurs = Fournisseur.objects.all().order_by('nom')

    # Sérialiser les désignations pour JavaScript
    designations_json = json.dumps([
        {'id': d.id, 'nom': d.nom} for d in designations
    ])

    return render(request, 'commande_bureau/commande_form_gestionnaire_bureau.html', {
        'form': CommandeBureauForm(instance=commande),
        'designations': designations,
        'designations_json': designations_json,
        'fournisseurs': fournisseurs,
        'commande': commande,
        'is_edit': True,
    })

def ajouter_commande(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        try:
            # Traitement des données du formulaire principal
            commande_data = {
                'mode_passation': request.POST.get('mode_passation'),
                'numero_commande': request.POST.get('numero_commande'),
                'fournisseur': request.POST.get('fournisseur'),
                'date_commande': request.POST.get('date_commande'),
                'date_reception': request.POST.get('date_reception'),
                'numero_facture': request.POST.get('numero_facture'),
                'duree_garantie_valeur': request.POST.get('duree_garantie_valeur'),
                'duree_garantie_unite': request.POST.get('duree_garantie_unite'),
            }

            commande_form = CommandeBureauForm(commande_data)

            if commande_form.is_valid():
                commande = commande_form.save()

                # Traitement des lignes de commande
                lignes_data = []
                i = 0
                while f'lignes-{i}-designation' in request.POST:
                    ligne_data = {
                        'designation_id': request.POST.get(f'lignes-{i}-designation'),
                        'description_id': request.POST.get(f'lignes-{i}-description'),
                        'quantite': request.POST.get(f'lignes-{i}-quantite'),
                        'prix_unitaire': request.POST.get(f'lignes-{i}-prix_unitaire'),
                    }

                    # Validation des données de ligne
                    if all([ligne_data['designation_id'], ligne_data['description_id'],
                            ligne_data['quantite'], ligne_data['prix_unitaire']]):
                        try:
                            LigneCommandeBureau.objects.create(
                                commande=commande,
                                designation_id=ligne_data['designation_id'],
                                description_id=ligne_data['description_id'],
                                quantite=int(ligne_data['quantite']),
                                prix_unitaire=float(ligne_data['prix_unitaire'])
                            )
                        except (ValueError, TypeError) as e:
                            logger.error(f"Erreur lors de la création de la ligne {i}: {e}")
                            messages.error(request, f"Erreur dans la ligne {i + 1}: données invalides")

                    i += 1

                if i > 0:  # Au moins une ligne créée
                    messages.success(request, f'Commande {commande.numero_commande} créée avec succès!')
                    return redirect('commandes_bureau:liste_commandes')
                else:
                    commande.delete()  # Supprimer la commande si aucune ligne n'est valide
                    messages.error(request, 'Aucune ligne de commande valide trouvée')
            else:
                messages.error(request, 'Erreur dans les données de la commande')

        except Exception as e:
            logger.error(f"Erreur lors de la création de la commande: {e}")
            messages.error(request, 'Erreur lors de la création de la commande')

    # GET request ou erreur dans POST
    designations = DesignationBureau.objects.all().order_by('nom')
    fournisseurs = Fournisseur.objects.all().order_by('nom')

    # Sérialiser les désignations pour JavaScript
    designations_json = json.dumps([
        {'id': d.id, 'nom': d.nom} for d in designations
    ])

    return render(request, 'commande_bureau/commande_form.html', {
        'form': CommandeBureauForm(),
        'designations': designations,
        'designations_json': designations_json,
        'fournisseurs': fournisseurs,
    })

def ajouter_commande_superadmin(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        try:
            commande_data = {
                'mode_passation': request.POST.get('mode_passation'),
                'numero_commande': request.POST.get('numero_commande'),
                'fournisseur': request.POST.get('fournisseur'),
                'date_commande': request.POST.get('date_commande'),
                'date_reception': request.POST.get('date_reception'),
                'numero_facture': request.POST.get('numero_facture'),
                'duree_garantie_valeur': request.POST.get('duree_garantie_valeur'),
                'duree_garantie_unite': request.POST.get('duree_garantie_unite'),
            }
            commande_form = CommandeBureauForm(commande_data)
            if commande_form.is_valid():
                commande = commande_form.save()
                i = 0
                while f'lignes-{i}-designation' in request.POST:
                    designation_id = request.POST.get(f'lignes-{i}-designation')
                    description_id = request.POST.get(f'lignes-{i}-description')
                    quantite = request.POST.get(f'lignes-{i}-quantite')
                    prix_unitaire = request.POST.get(f'lignes-{i}-prix_unitaire')
                    if all([designation_id, description_id, quantite, prix_unitaire]):
                        LigneCommandeBureau.objects.create(
                            commande=commande,
                            designation_id=designation_id,
                            description_id=description_id,
                            quantite=int(quantite),
                            prix_unitaire=float(prix_unitaire)
                        )
                    i += 1
                if i > 0:
                    messages.success(request, f'Commande {commande.numero_commande} créée avec succès!')
                    return redirect('commandes_bureau:liste_commandes_superadmin')
                else:
                    commande.delete()
                    messages.error(request, 'Aucune ligne de commande valide trouvée')
            else:
                messages.error(request, 'Erreur dans les données de la commande')
        except Exception as e:
            logger.error(f"Erreur lors de la création de la commande: {e}")
            messages.error(request, 'Erreur lors de la création de la commande')

    designations = DesignationBureau.objects.all().order_by('nom')
    fournisseurs = Fournisseur.objects.all().order_by('nom')
    designations_json = json.dumps([ { 'id': d.id, 'nom': d.nom } for d in designations ])
    return render(request, 'commande_bureau/commande_form_superadmin.html', {
        'form': CommandeBureauForm(),
        'designations': designations,
        'designations_json': designations_json,
        'fournisseurs': fournisseurs,
    })


@csrf_exempt
@require_http_methods(["POST"])
def ajouter_fournisseur(request):
    try:
        data = json.loads(request.body)
        nom = data.get('nom', '').strip()
        if_fiscal = data.get('if_fiscal', '').strip()
        ice = data.get('ice', '').strip()
        registre_commerce = data.get('registre_commerce', '').strip()
        adresse = data.get('adresse', '').strip()

        if not nom or not if_fiscal:
            return JsonResponse({'error': 'Le nom et l\'identifiant fiscal sont obligatoires'}, status=400)

        # Vérifier si le fournisseur existe déjà
        if Fournisseur.objects.filter(nom=nom).exists():
            return JsonResponse({'error': 'Un fournisseur avec ce nom existe déjà'}, status=400)

        if Fournisseur.objects.filter(if_fiscal=if_fiscal).exists():
            return JsonResponse({'error': 'Un fournisseur avec cet IF existe déjà'}, status=400)

        fournisseur = Fournisseur.objects.create(
            nom=nom,
            if_fiscal=if_fiscal,
            ice=ice or None,
            registre_commerce=registre_commerce or None,
            adresse=adresse or None
        )

        return JsonResponse({
            'id': fournisseur.id,
            'nom': fournisseur.nom,
            'success': True
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        logger.error(f"Erreur lors de l'ajout du fournisseur: {e}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def ajouter_designation(request):
    try:
        data = json.loads(request.body)
        nom = data.get('nom', '').strip()

        if not nom:
            return JsonResponse({'error': 'Le nom de la désignation est requis'}, status=400)

        # Vérifier si la désignation existe déjà
        if DesignationBureau.objects.filter(nom=nom).exists():
            return JsonResponse({'error': 'Cette désignation existe déjà'}, status=400)

        designation = DesignationBureau.objects.create(nom=nom)

        return JsonResponse({
            'id': designation.id,
            'nom': designation.nom,
            'success': True
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        logger.error(f"Erreur lors de l'ajout de la désignation: {e}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def ajouter_description(request):
    try:
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"Body reçu: {request.body}")
        data = json.loads(request.body)
        # Correction : accepter 'nom' comme champ (c'est ce que le JS envoie)
        nom = data.get('nom', '').strip()
        designation_id = data.get('designation_id')

        if not nom or not designation_id:
            return JsonResponse({'error': 'Le texte et la désignation sont requis'}, status=400)

        try:
            designation = DesignationBureau.objects.get(id=designation_id)
        except DesignationBureau.DoesNotExist:
            return JsonResponse({'error': 'Désignation introuvable'}, status=404)

        # Vérifier si la description existe déjà pour cette désignation
        if DescriptionBureau.objects.filter(designation=designation, nom=nom).exists():
            return JsonResponse({'error': 'Cette description existe déjà pour cette désignation'}, status=400)

        d = DescriptionBureau.objects.create(nom=nom, designation=designation)

        return JsonResponse({
            'id': d.id,
            'nom': d.nom,
            'designation_id': designation.id,
            'success': True
        })

    except json.JSONDecodeError:
        return JsonResponse({'error': 'Données JSON invalides'}, status=400)
    except Exception as e:
        logger.error(f"Erreur lors de l'ajout de la description: {e}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)


@require_http_methods(["GET"])
def get_descriptions(request, designation_id):
    try:
        try:
            designation = DesignationBureau.objects.get(id=designation_id)
        except DesignationBureau.DoesNotExist:
            return JsonResponse({'error': 'Désignation introuvable'}, status=404)

        descriptions = DescriptionBureau.objects.filter(designation=designation).order_by('nom')
        data = [{'id': d.id, 'nom': d.nom} for d in descriptions]

        return JsonResponse(data, safe=False)

    except Exception as e:
        logger.error(f"Erreur lors de la récupération des descriptions: {e}")
        return JsonResponse({'error': 'Erreur serveur'}, status=500)


def modifier_commande(request, pk):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    commande = get_object_or_404(CommandeBureau, pk=pk)

    if request.method == 'POST':
        try:
            # Traitement des données du formulaire principal
            commande_data = {
                'mode_passation': request.POST.get('mode_passation'),
                'numero_commande': request.POST.get('numero_commande'),
                'fournisseur': request.POST.get('fournisseur'),
                'date_commande': request.POST.get('date_commande'),
                'date_reception': request.POST.get('date_reception'),
                'numero_facture': request.POST.get('numero_facture'),
                'duree_garantie_valeur': request.POST.get('duree_garantie_valeur'),
                'duree_garantie_unite': request.POST.get('duree_garantie_unite'),
            }

            commande_form = CommandeBureauForm(commande_data, instance=commande)

            if commande_form.is_valid():
                commande = commande_form.save()

                # Récupérer les ids des lignes de commande dans le POST
                ids_post = set()
                i = 0
                while f'lignes-{i}-designation' in request.POST:
                    ligne_id = request.POST.get(f'lignes-{i}-id')  # Champ caché 'id' à ajouter dans le formset/template
                    if ligne_id:
                        ids_post.add(int(ligne_id))
                    i += 1

                # Supprimer uniquement les lignes absentes du POST
                for ligne in commande.lignes.all():
                    if ligne.id not in ids_post:
                        ligne.delete()

                # Mettre à jour ou créer les lignes présentes dans le POST
                i = 0
                while f'lignes-{i}-designation' in request.POST:
                    ligne_id = request.POST.get(f'lignes-{i}-id')
                    designation_id = request.POST.get(f'lignes-{i}-designation')
                    description_id = request.POST.get(f'lignes-{i}-description')
                    quantite = request.POST.get(f'lignes-{i}-quantite')
                    prix_unitaire = request.POST.get(f'lignes-{i}-prix_unitaire')
                    if all([designation_id, description_id, quantite, prix_unitaire]):
                        if ligne_id:
                            # Mise à jour de la ligne existante
                            try:
                                ligne = commande.lignes.get(id=ligne_id)
                                ligne.designation_id = designation_id
                                ligne.description_id = description_id
                                ligne.quantite = int(quantite)
                                ligne.prix_unitaire = float(prix_unitaire)
                                ligne.save()
                            except LigneCommandeBureau.DoesNotExist:
                                pass
                        else:
                            # Création d'une nouvelle ligne
                            LigneCommandeBureau.objects.create(
                                commande=commande,
                                designation_id=designation_id,
                                description_id=description_id,
                                quantite=int(quantite),
                                prix_unitaire=float(prix_unitaire)
                            )
                    i += 1

                if i > 0:
                    messages.success(request, f'Commande {commande.numero_commande} modifiée avec succès!')
                    return redirect('commandes_bureau:liste_commandes')
                else:
                    messages.error(request, 'Aucune ligne de commande valide trouvée')
            else:
                messages.error(request, 'Erreur dans les données de la commande')

        except Exception as e:
            logger.error(f"Erreur lors de la modification de la commande: {e}")
            messages.error(request, f'Erreur lors de la modification de la commande')

    # GET request ou erreur dans POST
    designations = DesignationBureau.objects.all().order_by('nom')
    fournisseurs = Fournisseur.objects.all().order_by('nom')

    # Sérialiser les désignations pour JavaScript
    designations_json = json.dumps([
        {'id': d.id, 'nom': d.nom} for d in designations
    ], ensure_ascii=False)

    # Sérialiser les lignes de commande existantes pour Alpine.js
    lignes_commande_json = json.dumps([
        (
            lambda descriptions: {
                'id': ligne.id,  # Ajout explicite de l'id
                'designationId': ligne.designation_id,
                'descriptionId': ligne.description_id,
                'quantite': ligne.quantite,
                'prix_unitaire': float(ligne.prix_unitaire),
                'descriptions': [
                    {'id': desc.id, 'nom': desc.nom}
                    for desc in descriptions
                ]
            }
        )(
            (lambda descs: descs + ([ligne.description] if ligne.description not in descs else []))(
                list(ligne.designation.descriptions.all())
            )
        )
        for ligne in commande.lignes.all()
    ], ensure_ascii=False)

    return render(request, 'commande_bureau/modifier_commande.html', {
        'form': CommandeBureauForm(instance=commande),
        'commande': commande,
        'fournisseurs': fournisseurs,
        'designations': designations,
        'designations_json': designations_json,
        'lignes_commande_json': lignes_commande_json,
        'is_edit': True,
    })

def modifier_commande_superadmin(request, pk):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    commande = get_object_or_404(CommandeBureau, pk=pk)
    if request.method == 'POST':
        try:
            commande_data = {
                'mode_passation': request.POST.get('mode_passation'),
                'numero_commande': request.POST.get('numero_commande'),
                'fournisseur': request.POST.get('fournisseur'),
                'date_commande': request.POST.get('date_commande'),
                'date_reception': request.POST.get('date_reception'),
                'numero_facture': request.POST.get('numero_facture'),
                'duree_garantie_valeur': request.POST.get('duree_garantie_valeur'),
                'duree_garantie_unite': request.POST.get('duree_garantie_unite'),
            }
            commande_form = CommandeBureauForm(commande_data, instance=commande)
            if commande_form.is_valid():
                commande = commande_form.save()
                ids_post = set()
                i = 0
                while f'lignes-{i}-designation' in request.POST:
                    ligne_id = request.POST.get(f'lignes-{i}-id')
                    if ligne_id:
                        ids_post.add(int(ligne_id))
                    i += 1
                for ligne in commande.lignes.all():
                    if ligne.id not in ids_post:
                        ligne.delete()
                i = 0
                while f'lignes-{i}-designation' in request.POST:
                    ligne_id = request.POST.get(f'lignes-{i}-id')
                    designation_id = request.POST.get(f'lignes-{i}-designation')
                    description_id = request.POST.get(f'lignes-{i}-description')
                    quantite = request.POST.get(f'lignes-{i}-quantite')
                    prix_unitaire = request.POST.get(f'lignes-{i}-prix_unitaire')
                    if all([designation_id, description_id, quantite, prix_unitaire]):
                        if ligne_id:
                            try:
                                ligne = commande.lignes.get(id=ligne_id)
                                ligne.designation_id = designation_id
                                ligne.description_id = description_id
                                ligne.quantite = int(quantite)
                                ligne.prix_unitaire = float(prix_unitaire)
                                ligne.save()
                            except LigneCommandeBureau.DoesNotExist:
                                pass
                        else:
                            LigneCommandeBureau.objects.create(
                                commande=commande,
                                designation_id=designation_id,
                                description_id=description_id,
                                quantite=int(quantite),
                                prix_unitaire=float(prix_unitaire)
                            )
                    i += 1
                if i > 0:
                    messages.success(request, f'Commande {commande.numero_commande} modifiée avec succès!')
                    return redirect('commandes_bureau:liste_commandes_superadmin')
                else:
                    messages.error(request, 'Aucune ligne de commande valide trouvée')
            else:
                messages.error(request, 'Erreur dans les données de la commande')
        except Exception as e:
            logger.error(f"Erreur lors de la modification de la commande: {e}")
            messages.error(request, f'Erreur lors de la modification de la commande')

    designations = DesignationBureau.objects.all().order_by('nom')
    fournisseurs = Fournisseur.objects.all().order_by('nom')
    designations_json = json.dumps([ { 'id': d.id, 'nom': d.nom } for d in designations ], ensure_ascii=False)
    lignes_commande_json = json.dumps([
        (lambda descriptions: {
            'id': ligne.id,
            'designationId': ligne.designation_id,
            'descriptionId': ligne.description_id,
            'quantite': ligne.quantite,
            'prix_unitaire': float(ligne.prix_unitaire),
            'descriptions': [ { 'id': desc.id, 'nom': desc.nom } for desc in descriptions ]
        })( (lambda descs: descs + ([ligne.description] if ligne.description not in descs else [])) ( list(ligne.designation.descriptions.all()) ) )
        for ligne in commande.lignes.all()
    ], ensure_ascii=False)
    return render(request, 'commande_bureau/modifier_commande_superadmin.html', {
        'form': CommandeBureauForm(instance=commande),
        'commande': commande,
        'fournisseurs': fournisseurs,
        'designations': designations,
        'designations_json': designations_json,
        'lignes_commande_json': lignes_commande_json,
        'is_edit': True,
    })


def supprimer_commande(request, pk):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    commande = get_object_or_404(CommandeBureau, pk=pk)
    commande.delete()
    messages.success(request, "Commande supprimée avec succès.")
    return redirect('commandes_bureau:liste_commandes')


def export_commandes_excel(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    commandes = CommandeBureau.objects.prefetch_related('lignes__designation', 'lignes__description', 'fournisseur')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Commandes Bureau"

    # En-têtes stylés (toutes les colonnes du tableau)
    headers = [
        "Mode de passation", "Numéro commande", "Fournisseur", "Date commande", "Date réception",
        "Numéro facture", "Durée garantie", "Désignation", "Description", "Quantité", "Prix unitaire"
    ]
    header_font = Font(bold=True, color="FFFFFF", size=14)
    # Dégradé bleu/violet (on simule avec deux couleurs alternées)
    header_fills = [PatternFill("solid", fgColor=color) for color in ["6366F1", "8B5CF6"]]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fills[(col_num-1) % 2]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin', color='6366F1'),
            right=Side(style='thin', color='6366F1'),
            top=Side(style='thin', color='6366F1'),
            bottom=Side(style='thin', color='6366F1'),
        )
    # Données : une ligne Excel par ligne de commande
    row_num = 2
    zebra_fills = [PatternFill("solid", fgColor="F3F4F6"), PatternFill("solid", fgColor="FFFFFF")]
    data_font = Font(size=12)
    for commande in commandes:
        lignes = list(commande.lignes.all())
        if lignes:
            for ligne in lignes:
                for col, value in enumerate([
                    commande.get_mode_passation_display() if hasattr(commande, 'get_mode_passation_display') else commande.mode_passation,
                    commande.numero_commande,
                    commande.fournisseur.nom if commande.fournisseur else "",
                    commande.date_commande.strftime("%d/%m/%Y") if commande.date_commande else "",
                    commande.date_reception.strftime("%d/%m/%Y") if commande.date_reception else "",
                    commande.numero_facture,
                    f"{commande.duree_garantie_valeur} {commande.get_duree_garantie_unite_display()}",
                    ligne.designation.nom if ligne.designation else "",
                    ligne.description.nom if ligne.description else "",
                    ligne.quantite,
                    ligne.prix_unitaire,
                ], 1):
                    cell = ws.cell(row=row_num, column=col, value=value)
                    cell.font = data_font
                    cell.fill = zebra_fills[(row_num-2) % 2]
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(
                        left=Side(style='thin', color='E5E7EB'),
                        right=Side(style='thin', color='E5E7EB'),
                        top=Side(style='thin', color='E5E7EB'),
                        bottom=Side(style='thin', color='E5E7EB'),
                    )
                row_num += 1
        else:
            for col, value in enumerate([
                commande.get_mode_passation_display() if hasattr(commande, 'get_mode_passation_display') else commande.mode_passation,
                commande.numero_commande,
                commande.fournisseur.nom if commande.fournisseur else "",
                commande.date_commande.strftime("%d/%m/%Y") if commande.date_commande else "",
                commande.date_reception.strftime("%d/%m/%Y") if commande.date_reception else "",
                commande.numero_facture,
                f"{commande.duree_garantie_valeur} {commande.get_duree_garantie_unite_display()}",
                "", "", "", ""
            ], 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.font = data_font
                cell.fill = zebra_fills[(row_num-2) % 2]
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=Side(style='thin', color='E5E7EB'),
                    right=Side(style='thin', color='E5E7EB'),
                    top=Side(style='thin', color='E5E7EB'),
                    bottom=Side(style='thin', color='E5E7EB'),
                )
            row_num += 1
    # Largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 4
    ws.row_dimensions[1].height = 28
    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=commandes_bureau.xlsx'
    wb.save(response)
    return response