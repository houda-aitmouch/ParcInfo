from django.shortcuts import render, redirect, get_object_or_404
from .models import MaterielInformatique
from .forms import MaterielInformatiqueForm
from django.http import JsonResponse
from apps.commande_informatique.models import LigneCommande
from django.core.exceptions import PermissionDenied
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required

def is_gestionnaire_ou_superadmin(user):
    return user.groups.filter(name__in=['Gestionnaire Informatique', 'Super Admin', 'Gestionnaire Bureau']).exists()

def is_gestionnaire_bureau(user):
    return user.groups.filter(name='Gestionnaire Bureau').exists()

def is_superadmin(user):
    return user.is_superuser


def liste_materiels(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    materiels = MaterielInformatique.objects.all()
    return render(request, 'materiel_informatique/liste_materiels.html', {'materiels': materiels})

def ajouter_materiel(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        form = MaterielInformatiqueForm(request.POST)
        if form.is_valid():
            materiel = form.save()
            return redirect('materiel_informatique:liste_materiels')
    else:
        form = MaterielInformatiqueForm()
    return render(request, 'materiel_informatique/ajouter_materiel.html', {'form': form})

def modifier_materiel(request, pk):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    materiel = get_object_or_404(MaterielInformatique, pk=pk)
    if request.method == 'POST':
        form = MaterielInformatiqueForm(request.POST, instance=materiel)
        if form.is_valid():
            materiel = form.save()
            return redirect('materiel_informatique:liste_materiels')
    else:
        form = MaterielInformatiqueForm(instance=materiel)
    return render(request, 'materiel_informatique/modifier_materiel.html', {
        'form': form,
        'is_edit': True,
        'materiel': materiel,
    })

def supprimer_materiel(request, pk):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    materiel = get_object_or_404(MaterielInformatique, pk=pk)
    if request.method == 'POST':
        materiel.delete()
        return redirect('materiel_informatique:liste_materiels')
    return render(request, 'materiel_informatique/confirmer_suppression.html', {'materiel': materiel})

def lignes_commande_par_commande(request, commande_id):
    lignes = LigneCommande.objects.filter(commande_id=commande_id).select_related('designation', 'description', 'commande__fournisseur', 'commande')
    data = [
        {
            'id': ligne.id,
            'designation': ligne.designation.nom,
            'designation_id': ligne.designation.id,
            'description': ligne.description.nom,
            'description_id': ligne.description.id,
            'prix_unitaire': str(ligne.prix_unitaire),
            'fournisseur': ligne.commande.fournisseur.nom if ligne.commande.fournisseur else '',
            'numero_facture': ligne.commande.numero_facture or '',
            'date_reception': ligne.commande.date_reception.strftime('%Y-%m-%d') if ligne.commande.date_reception else '',
            'duree_garantie_valeur': getattr(ligne.commande, 'duree_garantie_valeur', ''),
            'duree_garantie_unite': getattr(ligne.commande, 'duree_garantie_unite', ''),
        }
        for ligne in lignes
    ]
    return JsonResponse(data, safe=False)

def export_materiels_excel(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    materiels = MaterielInformatique.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Matériels"

    # En-tête
    headers = [
        "Commande", "Numéro de série", "Code inventaire", "Désignation", "Description", "Prix unitaire",
        "Fournisseur", "N° Facture", "Date service", "Date fin garantie", "Statut", "Utilisateur",
        "Lieu stockage", "Public", "Observation"
    ]
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Lignes
    for idx, mat in enumerate(materiels, start=2):
        ws.append([
            getattr(mat.ligne_commande.commande, 'numero_commande', ''),
            mat.numero_serie,
            mat.code_inventaire,
            getattr(mat.ligne_commande.designation, 'nom', ''),
            getattr(mat.ligne_commande.description, 'nom', ''),
            mat.ligne_commande.prix_unitaire,
            getattr(mat.ligne_commande.commande.fournisseur, 'nom', ''),
            getattr(mat.ligne_commande.commande, 'numero_facture', ''),
            mat.date_service_calculee.strftime('%d/%m/%Y') if mat.date_service_calculee else '',
            mat.date_fin_garantie_calculee.strftime('%d/%m/%Y') if mat.date_fin_garantie_calculee else '',
            dict(mat._meta.get_field('statut').choices).get(mat.statut, mat.statut),
            str(mat.utilisateur) if mat.utilisateur else '',
            dict(mat._meta.get_field('lieu_stockage').choices).get(mat.lieu_stockage, mat.lieu_stockage),
            "Oui" if mat.public else "Non",
            mat.observation,
        ])
        # Alternance de couleur de ligne
        fill = PatternFill(start_color="E0E7FF" if idx % 2 == 0 else "FFFFFF", end_color="E0E7FF" if idx % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if idx % 2 == 0:
                cell.fill = fill

    # Ajuste la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 3

    # Réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=liste_materiels.xlsx'
    wb.save(response)
    return response

@login_required
def mes_equipements_informatiques(request):
    equipements = MaterielInformatique.objects.filter(utilisateur=request.user, statut='affecte')
    
    # Déterminer le template de base selon le rôle de l'utilisateur
    user = request.user
    if user.is_superuser:
        base_template = 'dashboards/base_superadmin.html'
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        base_template = 'dashboards/base_gestionnaire_info.html'
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        base_template = 'dashboards/base_gestionnaire_bureau.html'
    elif user.groups.filter(name='Employés').exists():
        base_template = 'dashboards/base_employe.html'
    else:
        base_template = 'dashboards/base_employe.html'  # Par défaut pour les utilisateurs sans groupe
    
    context = {
        'equipements': equipements,
        'base_template': base_template
    }
    return render(request, 'materiel_informatique/mes_equipements_informatiques.html', context)

# ============================================================================
# VUES SUPERADMIN
# ============================================================================

def is_superadmin(user):
    """Vérifie si l'utilisateur est super admin"""
    return user.groups.filter(name='Super Admin').exists()

def is_gestionnaire_info(user):
    """Vérifie si l'utilisateur est gestionnaire info"""
    return user.groups.filter(name='Gestionnaire Informatique').exists()

def liste_materiels_superadmin(request):
    """Vue superadmin pour la liste des équipements informatiques"""
    if not is_superadmin(request.user):
        raise PermissionDenied
    
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Récupération des équipements avec pagination
    materiels_list = MaterielInformatique.objects.all().select_related(
        'ligne_commande__designation',
        'ligne_commande__description', 
        'ligne_commande__commande__fournisseur',
        'utilisateur'
    ).order_by('-id')
    
    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        materiels_list = materiels_list.filter(
            Q(numero_serie__icontains=search_query) |
            Q(code_inventaire__icontains=search_query) |
            Q(ligne_commande__designation__nom__icontains=search_query) |
            Q(ligne_commande__description__nom__icontains=search_query) |
            Q(ligne_commande__commande__fournisseur__nom__icontains=search_query) |
            Q(ligne_commande__commande__numero_facture__icontains=search_query) |
            Q(statut__icontains=search_query) |
            Q(utilisateur__username__icontains=search_query) |
            Q(utilisateur__first_name__icontains=search_query) |
            Q(utilisateur__last_name__icontains=search_query)
        )
    
    # Filtrage par statut
    status_filter = request.GET.get('status', '')
    if status_filter and status_filter != 'all':
        materiels_list = materiels_list.filter(statut=status_filter)
    
    # Pagination
    paginator = Paginator(materiels_list, 20)  # 20 équipements par page
    page_number = request.GET.get('page')
    materiels = paginator.get_page(page_number)
    
    context = {
        'materiels': materiels,
        'total': paginator.count,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    
    return render(request, 'materiel_informatique/liste_materiels_superadmin.html', context)

def liste_materiels_gestionnaire_info(request):
    """Vue gestionnaire info pour la liste des équipements informatiques"""
    if not (is_gestionnaire_info(request.user) or is_superadmin(request.user)):
        raise PermissionDenied
    
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Récupération des équipements avec pagination
    materiels_list = MaterielInformatique.objects.all().select_related(
        'ligne_commande__designation',
        'ligne_commande__description', 
        'ligne_commande__commande__fournisseur',
        'utilisateur'
    ).order_by('-id')
    
    # Recherche
    search_query = request.GET.get('search', '')
    if search_query:
        materiels_list = materiels_list.filter(
            Q(numero_serie__icontains=search_query) |
            Q(code_inventaire__icontains=search_query) |
            Q(ligne_commande__designation__nom__icontains=search_query) |
            Q(ligne_commande__description__nom__icontains=search_query) |
            Q(ligne_commande__commande__fournisseur__nom__icontains=search_query) |
            Q(ligne_commande__commande__numero_facture__icontains=search_query) |
            Q(statut__icontains=search_query) |
            Q(utilisateur__username__icontains=search_query) |
            Q(utilisateur__first_name__icontains=search_query) |
            Q(utilisateur__last_name__icontains=search_query)
        )
    
    # Filtrage par statut
    status_filter = request.GET.get('status', '')
    if status_filter and status_filter != 'all':
        materiels_list = materiels_list.filter(statut=status_filter)
    
    # Pagination
    paginator = Paginator(materiels_list, 20)  # 20 équipements par page
    page_number = request.GET.get('page')
    materiels = paginator.get_page(page_number)
    
    context = {
        'materiels': materiels,
        'total': paginator.count,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    
    return render(request, 'materiel_informatique/liste_materiels_gestionnaire_info.html', context)

def ajouter_materiel_superadmin(request):
    """Vue superadmin pour ajouter un équipement informatique"""
    if not is_superadmin(request.user):
        raise PermissionDenied
    
    if request.method == 'POST':
        form = MaterielInformatiqueForm(request.POST)
        if form.is_valid():
            materiel = form.save()
            return redirect('materiel_informatique:liste_materiels_superadmin')
    else:
        form = MaterielInformatiqueForm()
    
    context = {
        'form': form,
        'is_edit': False,
    }
    
    return render(request, 'materiel_informatique/ajouter_materiel_superadmin.html', context)

def modifier_materiel_superadmin(request, pk):
    """Vue superadmin pour modifier un équipement informatique"""
    if not is_superadmin(request.user):
        raise PermissionDenied
    
    materiel = get_object_or_404(MaterielInformatique, pk=pk)
    
    if request.method == 'POST':
        form = MaterielInformatiqueForm(request.POST, instance=materiel)
        if form.is_valid():
            materiel = form.save()
            return redirect('materiel_informatique:liste_materiels_superadmin')
    else:
        form = MaterielInformatiqueForm(instance=materiel)
    
    context = {
        'form': form,
        'materiel': materiel,
        'is_edit': True,
    }
    
    return render(request, 'materiel_informatique/modifier_materiel_superadmin.html', context)

def confirmer_suppression_superadmin(request, pk):
    """Vue superadmin pour confirmer la suppression d'un équipement"""
    if not is_superadmin(request.user):
        raise PermissionDenied
    
    materiel = get_object_or_404(MaterielInformatique, pk=pk)
    
    if request.method == 'POST':
        materiel.delete()
        return redirect('materiel_informatique:liste_materiels_superadmin')
    
    context = {
        'materiel': materiel,
    }
    
    return render(request, 'materiel_informatique/confirmer_suppression_superadmin.html', context)



def export_materiels_excel_superadmin(request):
    """Vue superadmin pour exporter les équipements en Excel"""
    if not is_superadmin(request.user):
        raise PermissionDenied
    
    materiels = MaterielInformatique.objects.all().select_related(
        'ligne_commande__designation',
        'ligne_commande__description', 
        'ligne_commande__commande__fournisseur',
        'utilisateur'
    ).order_by('-id')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Équipements Informatiques"

    # En-tête
    headers = [
        "Commande", "Numéro de série", "Code inventaire", "Désignation", "Description", "Prix unitaire",
        "Fournisseur", "N° Facture", "Date service", "Date fin garantie", "Statut", "Utilisateur",
        "Lieu stockage", "Public", "Observation"
    ]
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Lignes
    for idx, materiel in enumerate(materiels, start=2):
        ws.append([
            getattr(materiel.ligne_commande.commande, 'numero_commande', ''),
            materiel.numero_serie,
            materiel.code_inventaire,
            getattr(materiel.ligne_commande.designation, 'nom', ''),
            getattr(materiel.ligne_commande.description, 'nom', ''),
            materiel.ligne_commande.prix_unitaire,
            getattr(materiel.ligne_commande.commande.fournisseur, 'nom', ''),
            getattr(materiel.ligne_commande.commande, 'numero_facture', ''),
            materiel.date_service_calculee.strftime('%d/%m/%Y') if materiel.date_service_calculee else '',
            materiel.date_fin_garantie_calculee.strftime('%d/%m/%Y') if materiel.date_fin_garantie_calculee else '',
            dict(materiel._meta.get_field('statut').choices).get(materiel.statut, materiel.statut),
            str(materiel.utilisateur) if materiel.utilisateur else '',
            dict(materiel._meta.get_field('lieu_stockage').choices).get(materiel.lieu_stockage, materiel.lieu_stockage),
            "Oui" if materiel.public else "Non",
            materiel.observation,
        ])
        
        # Alternance de couleur de ligne
        fill = PatternFill(start_color="E0E7FF" if idx % 2 == 0 else "FFFFFF", end_color="E0E7FF" if idx % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if idx % 2 == 0:
                cell.fill = fill

    # Ajuste la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 3

    # Réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=equipements_informatiques_superadmin.xlsx'
    wb.save(response)
    return response

# ============================================================================
# VUES GESTIONNAIRE INFO (mêmes fonctionnalités, templates dédiés)
# ============================================================================

def ajouter_materiel_gestionnaire_info(request):
    """Vue gestionnaire info pour ajouter un équipement informatique"""
    if not is_gestionnaire_info(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        form = MaterielInformatiqueForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('materiel_informatique:liste_materiels_gestionnaire_info')
    else:
        form = MaterielInformatiqueForm()
    context = {
        'form': form,
        'is_edit': False,
    }
    return render(request, 'materiel_informatique/ajouter_materiel_gestionnaire_info.html', context)


def modifier_materiel_gestionnaire_info(request, pk: int):
    """Vue gestionnaire info pour modifier un équipement informatique"""
    if not is_gestionnaire_info(request.user):
        raise PermissionDenied
    materiel = get_object_or_404(MaterielInformatique, pk=pk)
    if request.method == 'POST':
        form = MaterielInformatiqueForm(request.POST, instance=materiel)
        if form.is_valid():
            form.save()
            return redirect('materiel_informatique:liste_materiels_gestionnaire_info')
    else:
        form = MaterielInformatiqueForm(instance=materiel)
    context = {
        'form': form,
        'materiel': materiel,
        'is_edit': True,
    }
    return render(request, 'materiel_informatique/modifier_materiel_gestionnaire_info.html', context)


def confirmer_suppression_gestionnaire_info(request, pk: int):
    """Vue gestionnaire info pour confirmer la suppression d'un équipement"""
    if not is_gestionnaire_info(request.user):
        raise PermissionDenied
    materiel = get_object_or_404(MaterielInformatique, pk=pk)
    if request.method == 'POST':
        materiel.delete()
        return redirect('materiel_informatique:liste_materiels_gestionnaire_info')
    context = {
        'materiel': materiel,
    }
    return render(request, 'materiel_informatique/confirmer_suppression_gestionnaire_info.html', context)

# ============================================================================
# VUES GESTIONNAIRE BUREAU (mêmes fonctionnalités, templates dédiés)
# ============================================================================

def liste_materiels_gestionnaire_bureau(request):
    """Vue gestionnaire bureau pour lister les équipements informatiques"""
    if not is_gestionnaire_bureau(request.user):
        raise PermissionDenied
    materiels = MaterielInformatique.objects.all()
    return render(request, 'materiel_informatique/liste_materiels_gestionnaire_bureau.html', {'materiels': materiels})

def ajouter_materiel_gestionnaire_bureau(request):
    """Vue gestionnaire bureau pour ajouter un équipement informatique"""
    if not is_gestionnaire_bureau(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        form = MaterielInformatiqueForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('materiel_informatique:liste_materiels_gestionnaire_bureau')
    else:
        form = MaterielInformatiqueForm()
    context = {
        'form': form,
        'is_edit': False,
    }
    return render(request, 'materiel_informatique/ajouter_materiel_gestionnaire_bureau.html', context)

def modifier_materiel_gestionnaire_bureau(request, pk: int):
    """Vue gestionnaire bureau pour modifier un équipement informatique"""
    if not is_gestionnaire_bureau(request.user):
        raise PermissionDenied
    materiel = get_object_or_404(MaterielInformatique, pk=pk)
    if request.method == 'POST':
        form = MaterielInformatiqueForm(request.POST, instance=materiel)
        if form.is_valid():
            form.save()
            return redirect('materiel_informatique:liste_materiels_gestionnaire_bureau')
    else:
        form = MaterielInformatiqueForm(instance=materiel)
    context = {
        'form': form,
        'materiel': materiel,
        'is_edit': True,
    }
    return render(request, 'materiel_informatique/modifier_materiel_gestionnaire_bureau.html', context)

def confirmer_suppression_gestionnaire_bureau(request, pk: int):
    """Vue gestionnaire bureau pour confirmer la suppression d'un équipement"""
    if not is_gestionnaire_bureau(request.user):
        raise PermissionDenied
    materiel = get_object_or_404(MaterielInformatique, pk=pk)
    if request.method == 'POST':
        materiel.delete()
        return redirect('materiel_informatique:liste_materiels_gestionnaire_bureau')
    context = {
        'materiel': materiel,
    }
    return render(request, 'materiel_informatique/confirmer_suppression_gestionnaire_bureau.html', context)
