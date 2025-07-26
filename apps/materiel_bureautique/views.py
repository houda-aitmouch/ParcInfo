from django.shortcuts import render, redirect, get_object_or_404
from .models import MaterielBureau
from .forms import MaterielBureauForm
from django.http import JsonResponse
from apps.commande_bureau.models import LigneCommandeBureau
from django.core.exceptions import PermissionDenied
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def is_gestionnaire_ou_superadmin(user):
    return user.groups.filter(name__in=['Gestionnaire Bureau', 'Super Admin']).exists()


def liste_materiels(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    materiels = MaterielBureau.objects.all()
    return render(request, 'materiel_bureautique/liste_materiels.html', {'materiels': materiels})

def ajouter_materiel(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        form = MaterielBureauForm(request.POST)
        if form.is_valid():
            # Récupérer la ligne de commande depuis les champs cachés
            designation_id = request.POST.get('designation')
            description_id = request.POST.get('description')
            
            if designation_id and description_id:
                try:
                    ligne_commande = LigneCommandeBureau.objects.get(
                        designation_id=designation_id,
                        description_id=description_id,
                        commande=form.cleaned_data['commande']
                    )
                    
                    # Compter le nombre de matériels existants pour cette ligne
                    materiels_existants = MaterielBureau.objects.filter(ligne_commande=ligne_commande).count()
                    
                    # Vérifier si on peut encore ajouter des matériels
                    if materiels_existants >= ligne_commande.quantite:
                        form.add_error(None, f"Impossible d'ajouter plus de matériels. Quantité commandée: {ligne_commande.quantite}, Matériels existants: {materiels_existants}")
                        return render(request, 'materiel_bureautique/ajouter_materiel.html', {'form': form})
                    
                    # Créer le matériel avec la ligne de commande trouvée
                    materiel = form.save(commit=False)
                    materiel.ligne_commande = ligne_commande
                    materiel.save()
                    return redirect('materiel_bureautique:liste_materiels')
                    
                except LigneCommandeBureau.DoesNotExist:
                    form.add_error(None, "Ligne de commande introuvable")
                    return render(request, 'materiel_bureautique/ajouter_materiel.html', {'form': form})
            else:
                form.add_error(None, "Veuillez sélectionner une désignation et une description")
                return render(request, 'materiel_bureautique/ajouter_materiel.html', {'form': form})
    else:
        form = MaterielBureauForm()
    return render(request, 'materiel_bureautique/ajouter_materiel.html', {'form': form})

def modifier_materiel(request, pk):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    materiel = get_object_or_404(MaterielBureau, pk=pk)
    if request.method == 'POST':
        form = MaterielBureauForm(request.POST, instance=materiel)
        if form.is_valid():
            # Récupérer la ligne de commande depuis les champs cachés
            designation_id = request.POST.get('designation')
            description_id = request.POST.get('description')
            
            if designation_id and description_id:
                try:
                    ligne_commande = LigneCommandeBureau.objects.get(
                        designation_id=designation_id,
                        description_id=description_id,
                        commande=form.cleaned_data['commande']
                    )
                    
                    # Compter le nombre de matériels existants pour cette ligne (excluant le matériel en cours de modification)
                    materiels_existants = MaterielBureau.objects.filter(ligne_commande=ligne_commande).exclude(pk=materiel.pk).count()
                    
                    # Vérifier si on peut encore ajouter des matériels
                    if materiels_existants >= ligne_commande.quantite:
                        form.add_error(None, f"Impossible d'ajouter plus de matériels. Quantité commandée: {ligne_commande.quantite}, Matériels existants: {materiels_existants}")
                        return render(request, 'materiel_bureautique/modifier_materiel.html', {
                            'form': form,
                            'is_edit': True,
                            'materiel': materiel,
                        })
                    
                    # Mettre à jour le matériel avec la ligne de commande trouvée
                    materiel = form.save(commit=False)
                    materiel.ligne_commande = ligne_commande
                    materiel.save()
                    return redirect('materiel_bureautique:liste_materiels')
                    
                except LigneCommandeBureau.DoesNotExist:
                    form.add_error(None, "Ligne de commande introuvable")
                    return render(request, 'materiel_bureautique/modifier_materiel.html', {
                        'form': form,
                        'is_edit': True,
                        'materiel': materiel,
                    })
            else:
                form.add_error(None, "Veuillez sélectionner une désignation et une description")
                return render(request, 'materiel_bureautique/modifier_materiel.html', {
                    'form': form,
                    'is_edit': True,
                    'materiel': materiel,
                })
    else:
        form = MaterielBureauForm(instance=materiel)
    return render(request, 'materiel_bureautique/modifier_materiel.html', {
        'form': form,
        'is_edit': True,
        'materiel': materiel,
    })

def supprimer_materiel(request, pk):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    materiel = get_object_or_404(MaterielBureau, pk=pk)
    if request.method == 'POST':
        materiel.delete()
        return redirect('materiel_bureautique:liste_materiels')
    return render(request, 'materiel_bureautique/confirmer_suppression.html', {'materiel': materiel})

def lignes_commande_par_commande(request, commande_id):
    lignes = LigneCommandeBureau.objects.filter(commande_id=commande_id).select_related('designation', 'description', 'commande__fournisseur', 'commande')
    data = []
    for ligne in lignes:
        # Compter le nombre de matériels existants pour cette ligne
        materiels_existants = MaterielBureau.objects.filter(ligne_commande=ligne).count()
        quantite_disponible = ligne.quantite - materiels_existants
        
        data.append({
            'id': ligne.id,
            'designation': ligne.designation.nom,
            'designation_id': ligne.designation.id,
            'description': ligne.description.nom,
            'description_id': ligne.description.id,
            'prix_unitaire': str(ligne.prix_unitaire),
            'fournisseur': ligne.commande.fournisseur.nom if ligne.commande.fournisseur else '',
            'numero_facture': ligne.commande.numero_facture or '',
            'date_reception': ligne.commande.date_reception.strftime('%Y-%m-%d') if ligne.commande.date_reception else '',
            'duree_garantie_valeur': getattr(ligne.commande, 'duree_garantie_valeur', ''),
            'duree_garantie_unite': getattr(ligne.commande, 'duree_garantie_unite', ''),
            'quantite_commande': ligne.quantite,
            'quantite_disponible': quantite_disponible,
            'materiels_existants': materiels_existants,
        })
    return JsonResponse(data, safe=False)

def export_materiels_excel(request):
    if not is_gestionnaire_ou_superadmin(request.user):
        raise PermissionDenied
    materiels = MaterielBureau.objects.all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Matériels Bureau"

    # En-tête
    headers = [
        "Commande", "Code inventaire", "Désignation", "Description", "Prix unitaire",
        "Fournisseur", "N° Facture", "Date service", "Date fin garantie", "Statut", "Utilisateur",
        "Lieu stockage", "Observation"
    ]
    ws.append(headers)

    # Styles
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Lignes
    for idx, mat in enumerate(materiels, start=2):
        ws.append([
            getattr(mat.ligne_commande.commande, 'numero_commande', ''),
            mat.code_inventaire,
            getattr(mat.ligne_commande.designation, 'nom', ''),
            getattr(mat.ligne_commande.description, 'nom', ''),
            mat.ligne_commande.prix_unitaire,
            getattr(mat.ligne_commande.commande.fournisseur, 'nom', ''),
            getattr(mat.ligne_commande.commande, 'numero_facture', ''),
            mat.date_service_calculee.strftime('%d/%m/%Y') if mat.date_service_calculee else '',
            mat.date_fin_garantie_calculee.strftime('%d/%m/%Y') if mat.date_fin_garantie_calculee else '',
            dict(mat._meta.get_field('statut').choices).get(mat.statut, mat.statut),
            str(mat.utilisateur) if mat.utilisateur else '',
            dict(mat._meta.get_field('lieu_stockage').choices).get(mat.lieu_stockage, mat.lieu_stockage) if mat.lieu_stockage else '',
            mat.observation,
        ])
        # Alternance de couleur de ligne
        fill = PatternFill(start_color="E0E7FF" if idx % 2 == 0 else "FFFFFF", end_color="E0E7FF" if idx % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            if idx % 2 == 0:
                cell.fill = fill

    # Ajuste la largeur des colonnes
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 3

    # Réponse HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=liste_materiels_bureau.xlsx'
    wb.save(response)
    return response
