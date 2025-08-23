from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.http import HttpResponse
from django.contrib.auth import get_user_model
from datetime import datetime
from openpyxl import Workbook
from .models import Livraison
from .forms import LivraisonForm, RechercheLivraisonForm, RechercheLivraisonBureauForm, NouvelleLivraisonForm

User = get_user_model()


@login_required
def liste_livraisons(request):
    """Affiche le tableau des livraisons avec toutes les commandes"""
    
    # Récupérer toutes les livraisons avec optimisations
    livraisons = Livraison.objects.select_related(
        'commande_informatique__fournisseur',
        'commande_bureau__fournisseur',
        'cree_par'
    ).prefetch_related(
        'commande_informatique__lignes__designation',
        'commande_informatique__lignes__description',
        'commande_bureau__lignes__designation',
        'commande_bureau__lignes__description'
    ).all()
    
    # Filtrage selon le type d'utilisateur
    user = request.user
    if user.is_superuser:
        # Superadmin voit tout
        pass
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        # Gestionnaire informatique voit seulement les commandes informatique
        livraisons = livraisons.filter(type_commande='informatique')
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        # Gestionnaire bureau voit seulement les commandes bureau
        livraisons = livraisons.filter(type_commande='bureau')
    else:
        # Utilisateurs normaux ne voient rien
        livraisons = Livraison.objects.none()
    
    # Filtres
    form = RechercheLivraisonForm(request.GET)
    if form.is_valid():
        search = form.cleaned_data.get('search')
        type_commande = form.cleaned_data.get('type_commande')
        statut_livraison = form.cleaned_data.get('statut_livraison')
        conforme = form.cleaned_data.get('conforme')
        pv_reception = form.cleaned_data.get('pv_reception')
        
        if search:
            livraisons = livraisons.filter(
                Q(numero_commande__icontains=search) |
                Q(commande_informatique__fournisseur__nom__icontains=search) |
                Q(commande_bureau__fournisseur__nom__icontains=search) |
                Q(notes__icontains=search)
            )
        
        if type_commande:
            livraisons = livraisons.filter(type_commande=type_commande)
        
        if statut_livraison:
            livraisons = livraisons.filter(statut_livraison=statut_livraison)
        
        if conforme and conforme in ['True', 'False']:
            livraisons = livraisons.filter(conforme=conforme == 'True')
        
        if pv_reception and pv_reception in ['True', 'False']:
            livraisons = livraisons.filter(pv_reception_recu=pv_reception == 'True')
    
    # Pagination
    paginator = Paginator(livraisons, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques
    total_livraisons = livraisons.count()
    livraisons_livrees = livraisons.filter(statut_livraison='livree').count()
    livraisons_conformes = livraisons.filter(conforme=True).count()
    pv_recus = livraisons.filter(pv_reception_recu=True).count()
    
    # Déterminer le dashboard approprié selon le type d'utilisateur
    dashboard_url = ''
    if user.is_superuser:
        dashboard_url = 'users:superadmin_dashboard'
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        dashboard_url = 'users:gestionnaire_info_dashboard'
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        dashboard_url = 'users:gestionnaire_bureau_dashboard'
    else:
        dashboard_url = 'users:employe_dashboard'
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_livraisons': total_livraisons,
        'livraisons_livrees': livraisons_livrees,
        'livraisons_conformes': livraisons_conformes,
        'pv_recus': pv_recus,
        'dashboard_url': dashboard_url,
    }
    
    return render(request, 'livraison/liste_livraisons.html', context)


@login_required
def detail_livraison(request, pk):
    """Détail d'une livraison avec les matériels"""
    livraison = get_object_or_404(Livraison.objects.select_related(
        'commande_informatique__fournisseur',
        'commande_bureau__fournisseur',
        'cree_par',
        'modifie_par'
    ).prefetch_related(
        'commande_informatique__lignes__designation',
        'commande_informatique__lignes__description',
        'commande_bureau__lignes__designation',
        'commande_bureau__lignes__description'
    ), pk=pk)
    
    # Vérifier les permissions
    user = request.user
    if not user.is_superuser:
        if livraison.type_commande == 'informatique' and not (user.groups.filter(name='Gestionnaire Informatique').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
        elif livraison.type_commande == 'bureau' and not (user.groups.filter(name='Gestionnaire Bureau').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
    
    # Calculer le total pour chaque matériel
    materiels_avec_total = []
    for materiel in livraison.materiels:
        materiel.total = materiel.prix_unitaire * materiel.quantite
        materiels_avec_total.append(materiel)
    
    context = {
        'livraison': livraison,
        'materiels': materiels_avec_total,
    }
    return render(request, 'livraison/detail_livraison.html', context)


@login_required
def nouvelle_livraison(request):
    """Créer une nouvelle livraison avec formulaire personnalisé"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Info: {user.groups.filter(name='Gestionnaire Informatique').exists()}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions
    if not user.is_superuser and not user.groups.filter(name='Gestionnaire Informatique').exists() and not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour créer une livraison.")
        return redirect('livraison:liste_livraisons')
    
    # Déterminer le type de commande selon l'utilisateur
    type_commande_initial = ''
    if user.is_superuser:
        type_commande_initial = ''
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        type_commande_initial = 'informatique'
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        type_commande_initial = 'bureau'
    
    if request.method == 'POST':
        form = NouvelleLivraisonForm(request.POST)
        if form.is_valid():
            # Traitement du formulaire
            commande_id = form.cleaned_data['commande_id']
            type_commande = form.cleaned_data['type_commande']
            date_livraison_prevue = form.cleaned_data['date_livraison_prevue']
            date_livraison_effective = form.cleaned_data['date_livraison_effective']
            statut_livraison = form.cleaned_data['statut_livraison']
            conforme = form.cleaned_data['conforme']
            pv_reception_recu = form.cleaned_data['pv_reception_recu']
            notes = form.cleaned_data['notes']
        
            # Utiliser le type de commande sélectionné dans le formulaire
            if type_commande == 'informatique':
                from apps.commande_informatique.models import Commande as CommandeInfo
                try:
                    commande_informatique = CommandeInfo.objects.get(id=commande_id)
                    commande_bureau = None
                except CommandeInfo.DoesNotExist:
                    messages.error(request, "Commande informatique introuvable.")
                    return redirect('livraison:nouvelle_livraison')
            elif type_commande == 'bureau':
                from apps.commande_bureau.models import CommandeBureau
                try:
                    commande_bureau = CommandeBureau.objects.get(id=commande_id)
                    commande_informatique = None
                except CommandeBureau.DoesNotExist:
                    messages.error(request, "Commande bureau introuvable.")
                    return redirect('livraison:nouvelle_livraison')
            else:
                messages.error(request, "Type de commande invalide.")
                return redirect('livraison:nouvelle_livraison')
            
            # Vérifier si une livraison existe déjà pour cette commande
            numero_commande = commande_informatique.numero_commande if commande_informatique else commande_bureau.numero_commande
            
            print(f"DEBUG - Vérification: numero_commande={numero_commande}, type_commande={type_commande}")
            
            livraison_existante = Livraison.objects.filter(numero_commande=numero_commande, type_commande=type_commande).first()
            if livraison_existante:
                print(f"DEBUG - Livraison existante trouvée: {livraison_existante}")
                messages.error(request, f'Une livraison existe déjà pour la commande {numero_commande}.')
                return redirect('livraison:detail_livraison', pk=livraison_existante.pk)
            else:
                print(f"DEBUG - Aucune livraison existante trouvée")
            
            # Créer la livraison
            print(f"DEBUG - Tentative de création: numero_commande={numero_commande}, type_commande={type_commande}")
            try:
                livraison = Livraison.objects.create(
                    numero_commande=numero_commande,
                    type_commande=type_commande,
                    commande_informatique=commande_informatique,
                    commande_bureau=commande_bureau,
                    date_livraison_prevue=date_livraison_prevue,
                    date_livraison_effective=date_livraison_effective or None,
                    statut_livraison=statut_livraison,
                    conforme=conforme,
                    pv_reception_recu=pv_reception_recu,
                    notes=notes,
                    cree_par=user
                )
                
                print(f"DEBUG - Livraison créée avec succès: {livraison}")
                messages.success(request, f'Livraison pour la commande {livraison.numero_commande} créée avec succès.')
                return redirect('livraison:detail_livraison', pk=livraison.pk)
            except Exception as e:
                print(f"DEBUG - Erreur lors de la création: {str(e)}")
                messages.error(request, f'Erreur lors de la création de la livraison : {str(e)}')
                return redirect('livraison:nouvelle_livraison')
        else:
            # Formulaire invalide - afficher les erreurs
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        # GET request - afficher le formulaire
        form = NouvelleLivraisonForm(initial={'type_commande': type_commande_initial})
    
    context = {
        'form': form,
        'type_commande_initial': type_commande_initial
    }
    return render(request, 'livraison/form_livraison.html', context)


@login_required
def modifier_livraison(request, pk):
    """Modifier une livraison"""
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier les permissions
    user = request.user
    if not user.is_superuser:
        if livraison.type_commande == 'informatique' and not (user.groups.filter(name='Gestionnaire Informatique').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
        elif livraison.type_commande == 'bureau' and not (user.groups.filter(name='Gestionnaire Bureau').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
    
    if request.method == 'POST':
        from .forms import ModifierLivraisonForm
        form = ModifierLivraisonForm(request.POST)
        
        if form.is_valid():
            # Mettre à jour la livraison avec les données du formulaire
            livraison.date_livraison_prevue = form.cleaned_data['date_livraison_prevue']
            livraison.date_livraison_effective = form.cleaned_data['date_livraison_effective'] or None
            livraison.statut_livraison = form.cleaned_data['statut_livraison']
            livraison.conforme = form.cleaned_data['conforme']
            livraison.pv_reception_recu = form.cleaned_data['pv_reception_recu']
            livraison.notes = form.cleaned_data['notes']
            livraison.modifie_par = user
            livraison.save()
            
            messages.success(request, f'Livraison {livraison.numero_commande} modifiée avec succès.')
            return redirect('livraison:detail_livraison', pk=livraison.pk)
        else:
            # Afficher les erreurs spécifiques pour le debug
            print("DEBUG - Erreurs de formulaire:", form.errors)
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        # Pré-remplir le formulaire avec les données existantes
        from .forms import ModifierLivraisonForm
        initial_data = {
            'type_commande': livraison.type_commande,
            'commande_id': livraison.commande_informatique.id if livraison.commande_informatique else livraison.commande_bureau.id,
            'date_livraison_prevue': livraison.date_livraison_prevue.strftime('%Y-%m-%d') if livraison.date_livraison_prevue else None,
            'date_livraison_effective': livraison.date_livraison_effective.strftime('%Y-%m-%d') if livraison.date_livraison_effective else None,
            'statut_livraison': livraison.statut_livraison,
            'conforme': livraison.conforme,
            'pv_reception_recu': livraison.pv_reception_recu,
            'notes': livraison.notes,
        }
        
        # Debug: afficher les valeurs initiales
        print("DEBUG - Valeurs initiales:", initial_data)
        
        form = ModifierLivraisonForm(initial=initial_data)
    
    # Calculer le total pour chaque matériel (comme dans detail_livraison)
    materiels_avec_total = []
    for materiel in livraison.materiels:
        materiel.total = materiel.prix_unitaire * materiel.quantite
        materiels_avec_total.append(materiel)
    
    context = {
        'form': form,
        'livraison': livraison,
        'materiels': materiels_avec_total,  # Passer les matériels avec total calculé
        'titre': f'Modifier la livraison {livraison.numero_commande}',
        'is_edit': True  # Flag pour indiquer qu'on est en mode modification
    }
    return render(request, 'livraison/form_livraison.html', context)


@login_required
def supprimer_livraison(request, pk):
    """Supprimer une livraison"""
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier les permissions
    user = request.user
    if not user.is_superuser:
        if livraison.type_commande == 'informatique' and not (user.groups.filter(name='Gestionnaire Informatique').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
        elif livraison.type_commande == 'bureau' and not (user.groups.filter(name='Gestionnaire Bureau').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
    
    if request.method == 'POST':
        numero = livraison.numero_commande
        livraison.delete()
        messages.success(request, f'Livraison {numero} supprimée avec succès.')
        return redirect('livraison:liste_livraisons')
    
    context = {
        'livraison': livraison
    }
    return render(request, 'livraison/confirmer_suppression.html', context)


@login_required
def marquer_livree(request, pk):
    """Marquer une livraison comme livrée"""
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier les permissions
    user = request.user
    if not user.is_superuser:
        if livraison.type_commande == 'informatique' and not (user.groups.filter(name='Gestionnaire Informatique').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
        elif livraison.type_commande == 'bureau' and not (user.groups.filter(name='Gestionnaire Bureau').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
    
    if request.method == 'POST':
        livraison.statut_livraison = 'livree'
        livraison.modifie_par = user
        livraison.save()
        
        messages.success(request, f'Livraison {livraison.numero_commande} marquée comme livrée.')
        return redirect('livraison:detail_livraison', pk=pk)
    
    context = {
        'livraison': livraison
    }
    return render(request, 'livraison/marquer_livree.html', context)


@login_required
def valider_pv_reception(request, pk):
    """Valider le PV de réception"""
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier les permissions
    user = request.user
    if not user.is_superuser:
        if livraison.type_commande == 'informatique' and not (user.groups.filter(name='Gestionnaire Informatique').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
        elif livraison.type_commande == 'bureau' and not (user.groups.filter(name='Gestionnaire Bureau').exists()):
            messages.error(request, "Vous n'avez pas accès à cette livraison.")
            return redirect('livraison:liste_livraisons')
    
    if request.method == 'POST':
        livraison.pv_reception_recu = True
        livraison.modifie_par = user
        livraison.save()
        
        messages.success(request, f'PV de réception validé pour la livraison {livraison.numero_commande}.')
        return redirect('livraison:detail_livraison', pk=pk)
    
    context = {
        'livraison': livraison
    }
    return render(request, 'livraison/valider_pv_reception.html', context)


@login_required
def rapports_livraison(request):
    """Rapports et statistiques de livraison"""
    
    # Récupérer les livraisons selon les permissions
    user = request.user
    if user.is_superuser:
        livraisons = Livraison.objects.all()
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        livraisons = Livraison.objects.filter(type_commande='informatique')
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        livraisons = Livraison.objects.filter(type_commande='bureau')
    else:
        livraisons = Livraison.objects.none()
    
    # Statistiques générales
    total_livraisons = livraisons.count()
    livraisons_livrees = livraisons.filter(statut_livraison='livree').count()
    livraisons_conformes = livraisons.filter(conforme=True).count()
    pv_recus = livraisons.filter(pv_reception_recu=True).count()
    
    # Par type de commande
    livraisons_informatique = livraisons.filter(type_commande='informatique').count()
    livraisons_bureau = livraisons.filter(type_commande='bureau').count()
    
    # Par statut
    livraisons_par_statut = livraisons.values('statut_livraison').annotate(
        count=Count('id')
    ).order_by('statut_livraison')
    
    # Par fournisseur
    livraisons_par_fournisseur = livraisons.values(
        'commande_informatique__fournisseur__nom'
    ).annotate(
        count=Count('id')
    ).filter(
        commande_informatique__fournisseur__isnull=False
    ).order_by('-count')[:10]
    
    context = {
        'total_livraisons': total_livraisons,
        'livraisons_livrees': livraisons_livrees,
        'livraisons_conformes': livraisons_conformes,
        'pv_recus': pv_recus,
        'livraisons_informatique': livraisons_informatique,
        'livraisons_bureau': livraisons_bureau,
        'livraisons_par_statut': livraisons_par_statut,
        'livraisons_par_fournisseur': livraisons_par_fournisseur,
    }
    
    return render(request, 'livraison/rapports.html', context)


# Vues API pour le formulaire dynamique
@login_required
def api_commandes(request):
    """API pour récupérer les commandes selon le type"""
    user = request.user
    # Accepter les deux paramètres pour la compatibilité
    type_commande = request.GET.get('type') or request.GET.get('type_commande')
    
    # Vérifier les permissions (le superadmin a accès à tout)
    if not user.is_superuser:
        if type_commande == 'informatique' and not (user.groups.filter(name='Gestionnaire Informatique').exists()):
            return JsonResponse({'success': False, 'error': 'Permission refusée'}, status=403)
        elif type_commande == 'bureau' and not (user.groups.filter(name='Gestionnaire Bureau').exists()):
            return JsonResponse({'success': False, 'error': 'Permission refusée'}, status=403)
    
    commandes = []
    
    if type_commande == 'informatique':
        from apps.commande_informatique.models import Commande
        commandes_objs = Commande.objects.select_related('fournisseur').all()
        commandes = [
            {
                'id': cmd.id,
                'numero_commande': cmd.numero_commande,
                'fournisseur': cmd.fournisseur.nom if cmd.fournisseur else 'Non spécifié',
                'date_commande': cmd.date_commande.strftime('%d/%m/%Y') if cmd.date_commande else 'Non spécifiée'
            }
            for cmd in commandes_objs
        ]
    elif type_commande == 'bureau':
        from apps.commande_bureau.models import CommandeBureau
        commandes_objs = CommandeBureau.objects.select_related('fournisseur').all()
        commandes = [
            {
                'id': cmd.id,
                'numero_commande': cmd.numero_commande,
                'fournisseur': cmd.fournisseur.nom if cmd.fournisseur else 'Non spécifié',
                'date_commande': cmd.date_commande.strftime('%d/%m/%Y') if cmd.date_commande else 'Non spécifiée'
            }
            for cmd in commandes_objs
        ]
    
    return JsonResponse({'success': True, 'commandes': commandes})


@login_required
def api_commande_details(request, commande_id):
    """API pour récupérer les détails d'une commande"""
    user = request.user
    # Accepter les deux paramètres pour la compatibilité
    type_commande = request.GET.get('type') or request.GET.get('type_commande')
    
    # Convertir commande_id en integer
    try:
        commande_id = int(commande_id)
    except (ValueError, TypeError):
        print(f"DEBUG - ID de commande invalide: {commande_id}")
        return JsonResponse({'success': False, 'error': 'ID de commande invalide'}, status=400)
    
    print(f"DEBUG - API détails: commande_id={commande_id} (type: {type(commande_id)}), type_commande={type_commande}")
    print(f"DEBUG - Utilisateur: {user.username}")
    
    # Vérifier les permissions (le superadmin a accès à tout)
    if not user.is_superuser:
        if type_commande == 'informatique' and not (user.groups.filter(name='Gestionnaire Informatique').exists()):
            print(f"DEBUG - Permission refusée pour informatique")
            return JsonResponse({'success': False, 'error': 'Permission refusée'}, status=403)
        elif type_commande == 'bureau' and not (user.groups.filter(name='Gestionnaire Bureau').exists()):
            print(f"DEBUG - Permission refusée pour bureau")
            return JsonResponse({'success': False, 'error': 'Permission refusée'}, status=403)
    
    try:
        if type_commande == 'informatique':
            from apps.commande_informatique.models import Commande
            print(f"DEBUG - Recherche commande informatique ID: {commande_id}")
            commande = Commande.objects.select_related('fournisseur').prefetch_related('lignes__designation', 'lignes__description').get(id=commande_id)
            print(f"DEBUG - Commande trouvée: {commande.numero_commande}")
            
            materiels = []
            montant_total = 0
            nombre_articles = 0
            
            for ligne in commande.lignes.all():
                prix_total_ligne = ligne.quantite * ligne.prix_unitaire
                montant_total += prix_total_ligne
                nombre_articles += ligne.quantite
                
                materiels.append({
                    'id': ligne.id,
                    'designation': ligne.designation.nom if ligne.designation else 'Non spécifié',
                    'description': ligne.description.nom if ligne.description else '',
                    'quantite': ligne.quantite,
                    'prix_unitaire': f"{ligne.prix_unitaire:.2f}",
                    'prix_total': f"{prix_total_ligne:.2f}"
                })
            
            print(f"DEBUG - Nombre de lignes: {len(materiels)}")
            
        elif type_commande == 'bureau':
            from apps.commande_bureau.models import CommandeBureau
            commande = CommandeBureau.objects.select_related('fournisseur').prefetch_related('lignes__designation', 'lignes__description').get(id=commande_id)
            
            materiels = []
            montant_total = 0
            nombre_articles = 0
            
            for ligne in commande.lignes.all():
                prix_total_ligne = ligne.quantite * ligne.prix_unitaire
                montant_total += prix_total_ligne
                nombre_articles += ligne.quantite
                
                materiels.append({
                    'id': ligne.id,
                    'designation': ligne.designation.nom if ligne.designation else 'Non spécifié',
                    'description': ligne.description.nom if ligne.description else '',
                    'quantite': ligne.quantite,
                    'prix_unitaire': f"{ligne.prix_unitaire:.2f}",
                    'prix_total': f"{prix_total_ligne:.2f}"
                })
        else:
            print(f"DEBUG - Type de commande invalide: {type_commande}")
            return JsonResponse({'success': False, 'error': 'Type de commande invalide'}, status=400)
        
        response_data = {
            'success': True,
            'commande': {
                'fournisseur': commande.fournisseur.nom if commande.fournisseur else 'Non spécifié',
                'numero_commande': commande.numero_commande,
                'montant_total': f"{montant_total:.2f}",
                'nombre_articles': nombre_articles
            },
            'materiels': materiels
        }
        
        print(f"DEBUG - Réponse envoyée: {response_data}")
        return JsonResponse(response_data)
        
    except Exception as e:
        print(f"DEBUG - Exception: {str(e)}")
        import traceback
        print(f"DEBUG - Traceback: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'error': f'Erreur: {str(e)}'}, status=404)


@login_required
def export_livraisons_excel(request):
    """Exporter les livraisons (vue standard) en Excel"""
    # Filtrer selon permissions comme dans liste_livraisons
    livraisons = Livraison.objects.all()
    user = request.user
    if user.is_superuser:
        pass
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        livraisons = livraisons.filter(type_commande='informatique')
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        livraisons = livraisons.filter(type_commande='bureau')
    else:
        livraisons = Livraison.objects.none()

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Livraisons"

    headers = [
        "N° Commande", "Type", "Fournisseur", "Date prévue", "Date effective",
        "Statut", "Conforme", "PV reçu"
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    zebra_fills = [PatternFill("solid", fgColor="F9FAFB"), PatternFill("solid", fgColor="FFFFFF")]
    for liv in livraisons:
        values = [
            liv.numero_commande,
            liv.get_type_commande_display(),
            getattr(liv.fournisseur, 'nom', '') if hasattr(liv, 'fournisseur') else '',
            liv.date_livraison_prevue.strftime('%d/%m/%Y') if liv.date_livraison_prevue else '',
            liv.date_livraison_effective.strftime('%d/%m/%Y') if liv.date_livraison_effective else '',
            liv.get_statut_livraison_display(),
            'Oui' if liv.conforme else 'Non',
            'Oui' if liv.pv_reception_recu else 'Non'
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = zebra_fills[(row-2) % 2]
        row += 1

    # Ajuster largeur
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 4

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=livraisons.xlsx'
    wb.save(response)
    return response


# ===== Superadmin variants (render superadmin templates) =====
@login_required
def liste_livraisons_superadmin(request):
    # Reuse liste_livraisons logic, render superadmin template
    # Récupérations identiques
    livraisons = Livraison.objects.select_related(
        'commande_informatique__fournisseur',
        'commande_bureau__fournisseur',
        'cree_par'
    ).prefetch_related(
        'commande_informatique__lignes__designation',
        'commande_informatique__lignes__description',
        'commande_bureau__lignes__designation',
        'commande_bureau__lignes__description'
    ).all()
    user = request.user
    if user.is_superuser:
        pass
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        livraisons = livraisons.filter(type_commande='informatique')
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        livraisons = livraisons.filter(type_commande='bureau')
    else:
        livraisons = Livraison.objects.none()

    form = RechercheLivraisonForm(request.GET)
    if form.is_valid():
        search = form.cleaned_data.get('search')
        type_commande = form.cleaned_data.get('type_commande')
        statut_livraison = form.cleaned_data.get('statut_livraison')
        conforme = form.cleaned_data.get('conforme')
        pv_reception = form.cleaned_data.get('pv_reception')
        if search:
            livraisons = livraisons.filter(
                Q(numero_commande__icontains=search) |
                Q(commande_informatique__fournisseur__nom__icontains=search) |
                Q(commande_bureau__fournisseur__nom__icontains=search) |
                Q(notes__icontains=search)
            )
        if type_commande:
            livraisons = livraisons.filter(type_commande=type_commande)
        if statut_livraison:
            livraisons = livraisons.filter(statut_livraison=statut_livraison)
        if conforme and conforme in ['True', 'False']:
            livraisons = livraisons.filter(conforme=conforme == 'True')
        if pv_reception and pv_reception in ['True', 'False']:
            livraisons = livraisons.filter(pv_reception_recu=pv_reception == 'True')

    paginator = Paginator(livraisons, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_livraisons': livraisons.count(),
        'livraisons_livrees': livraisons.filter(statut_livraison='livree').count(),
        'livraisons_conformes': livraisons.filter(conforme=True).count(),
        'pv_recus': livraisons.filter(pv_reception_recu=True).count(),
    }
    return render(request, 'livraison/liste_livraisons_superadmin.html', context)


@login_required
def liste_livraisons_gestionnaire_info(request):
    # Reuse liste_livraisons logic, render gestionnaire info template
    # Récupérations identiques
    livraisons = Livraison.objects.select_related(
        'commande_informatique__fournisseur',
        'commande_bureau__fournisseur',
        'cree_par'
    ).prefetch_related(
        'commande_informatique__lignes__designation',
        'commande_informatique__lignes__description',
        'commande_bureau__lignes__designation',
        'commande_bureau__lignes__description'
    ).all()
    user = request.user
    if user.is_superuser:
        pass
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        livraisons = livraisons.filter(type_commande='informatique')
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        livraisons = livraisons.filter(type_commande='bureau')
    else:
        livraisons = Livraison.objects.none()

    # Pour le gestionnaire info, forcer le type_commande à informatique par défaut
    get_data = request.GET.copy()
    if "type_commande" not in get_data:
        get_data["type_commande"] = "informatique"
    
    form = RechercheLivraisonForm(get_data)
    if form.is_valid():
        search = form.cleaned_data.get('search')
        type_commande = form.cleaned_data.get('type_commande')
        statut_livraison = form.cleaned_data.get('statut_livraison')
        conforme = form.cleaned_data.get('conforme')
        pv_reception = form.cleaned_data.get('pv_reception')
        if search:
            livraisons = livraisons.filter(
                Q(numero_commande__icontains=search) |
                Q(commande_informatique__fournisseur__nom__icontains=search) |
                Q(commande_bureau__fournisseur__nom__icontains=search) |
                Q(notes__icontains=search)
            )
        # Pour le gestionnaire info, toujours filtrer sur informatique
        livraisons = livraisons.filter(type_commande='informatique')
        if statut_livraison:
            livraisons = livraisons.filter(statut_livraison=statut_livraison)
        if conforme and conforme in ['True', 'False']:
            livraisons = livraisons.filter(conforme=conforme == 'True')
        if pv_reception and pv_reception in ['True', 'False']:
            livraisons = livraisons.filter(pv_reception_recu=pv_reception == 'True')

    paginator = Paginator(livraisons, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_livraisons': livraisons.count(),
        'livraisons_livrees': livraisons.filter(statut_livraison='livree').count(),
        'livraisons_conformes': livraisons.filter(conforme=True).count(),
        'pv_recus': livraisons.filter(pv_reception_recu=True).count(),
    }
    return render(request, 'livraison/liste_livraisons_gestionnaire_info.html', context)


@login_required
def detail_livraison_superadmin(request, pk):
    livraison = get_object_or_404(Livraison.objects.select_related(
        'commande_informatique__fournisseur',
        'commande_bureau__fournisseur',
        'cree_par',
        'modifie_par'
    ).prefetch_related(
        'commande_informatique__lignes__designation',
        'commande_informatique__lignes__description',
        'commande_bureau__lignes__designation',
        'commande_bureau__lignes__description'
    ), pk=pk)
    materiels_avec_total = []
    for materiel in livraison.materiels:
        materiel.total = materiel.prix_unitaire * materiel.quantite
        materiels_avec_total.append(materiel)
    return render(request, 'livraison/detail_livraison_superadmin.html', {
        'livraison': livraison,
        'materiels': materiels_avec_total,
    })


@login_required
def nouvelle_livraison_superadmin(request):
    # Same as nouvelle_livraison but renders superadmin template and redirects to superadmin detail
    user = request.user
    if not user.is_superuser and not (user.groups.filter(name='Gestionnaire Informatique').exists()) and not (user.groups.filter(name='Gestionnaire Bureau').exists()):
        messages.error(request, "Vous n'avez pas les permissions pour créer une livraison.")
        return redirect('livraison:liste_livraisons_superadmin')
    type_commande_initial = ''
    if user.is_superuser:
        type_commande_initial = ''
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        type_commande_initial = 'informatique'
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        type_commande_initial = 'bureau'
    if request.method == 'POST':
        form = NouvelleLivraisonForm(request.POST)
        if form.is_valid():
            commande_id = form.cleaned_data['commande_id']
            type_commande = form.cleaned_data['type_commande']
            date_livraison_prevue = form.cleaned_data['date_livraison_prevue']
            date_livraison_effective = form.cleaned_data['date_livraison_effective']
            statut_livraison = form.cleaned_data['statut_livraison']
            conforme = form.cleaned_data['conforme']
            pv_reception_recu = form.cleaned_data['pv_reception_recu']
            notes = form.cleaned_data['notes']
            if type_commande == 'informatique':
                from apps.commande_informatique.models import Commande as CommandeInfo
                try:
                    commande_informatique = CommandeInfo.objects.get(id=commande_id)
                    commande_bureau = None
                except CommandeInfo.DoesNotExist:
                    messages.error(request, "Commande informatique introuvable.")
                    return redirect('livraison:nouvelle_livraison_superadmin')
            elif type_commande == 'bureau':
                from apps.commande_bureau.models import CommandeBureau
                try:
                    commande_bureau = CommandeBureau.objects.get(id=commande_id)
                    commande_informatique = None
                except CommandeBureau.DoesNotExist:
                    messages.error(request, "Commande bureau introuvable.")
                    return redirect('livraison:nouvelle_livraison_superadmin')
            else:
                messages.error(request, "Type de commande invalide.")
                return redirect('livraison:nouvelle_livraison_superadmin')
            numero_commande = commande_informatique.numero_commande if commande_informatique else commande_bureau.numero_commande
            livraison_existante = Livraison.objects.filter(numero_commande=numero_commande, type_commande=type_commande).first()
            if livraison_existante:
                messages.error(request, f'Une livraison existe déjà pour la commande {numero_commande}.')
                return redirect('livraison:detail_livraison_superadmin', pk=livraison_existante.pk)
            try:
                livraison = Livraison.objects.create(
                    numero_commande=numero_commande,
                    type_commande=type_commande,
                    commande_informatique=commande_informatique,
                    commande_bureau=commande_bureau,
                    date_livraison_prevue=date_livraison_prevue,
                    date_livraison_effective=date_livraison_effective or None,
                    statut_livraison=statut_livraison,
                    conforme=conforme,
                    pv_reception_recu=pv_reception_recu,
                    notes=notes,
                    cree_par=user
                )
                messages.success(request, f'Livraison pour la commande {livraison.numero_commande} créée avec succès.')
                return redirect('livraison:detail_livraison_superadmin', pk=livraison.pk)
            except Exception as e:
                messages.error(request, f'Erreur lors de la création de la livraison : {str(e)}')
                return redirect('livraison:nouvelle_livraison_superadmin')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        form = NouvelleLivraisonForm(initial={'type_commande': type_commande_initial})
    return render(request, 'livraison/form_livraison_superadmin.html', { 'form': form, 'type_commande_initial': type_commande_initial })


@login_required
def modifier_livraison_superadmin(request, pk):
    livraison = get_object_or_404(Livraison, pk=pk)
    user = request.user
    if request.method == 'POST':
        from .forms import ModifierLivraisonForm
        form = ModifierLivraisonForm(request.POST)
        if form.is_valid():
            livraison.date_livraison_prevue = form.cleaned_data['date_livraison_prevue']
            livraison.date_livraison_effective = form.cleaned_data['date_livraison_effective'] or None
            livraison.statut_livraison = form.cleaned_data['statut_livraison']
            livraison.conforme = form.cleaned_data['conforme']
            livraison.pv_reception_recu = form.cleaned_data['pv_reception_recu']
            livraison.notes = form.cleaned_data['notes']
            livraison.modifie_par = user
            livraison.save()
            messages.success(request, f'Livraison {livraison.numero_commande} modifiée avec succès.')
            return redirect('livraison:detail_livraison_superadmin', pk=livraison.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        from .forms import ModifierLivraisonForm
        initial_data = {
            'type_commande': livraison.type_commande,
            'commande_id': livraison.commande_informatique.id if livraison.commande_informatique else livraison.commande_bureau.id,
            'date_livraison_prevue': livraison.date_livraison_prevue.strftime('%Y-%m-%d') if livraison.date_livraison_prevue else None,
            'date_livraison_effective': livraison.date_livraison_effective.strftime('%Y-%m-%d') if livraison.date_livraison_effective else None,
            'statut_livraison': livraison.statut_livraison,
            'conforme': livraison.conforme,
            'pv_reception_recu': livraison.pv_reception_recu,
            'notes': livraison.notes,
        }
        form = ModifierLivraisonForm(initial=initial_data)
    # Matériels totaux
    materiels_avec_total = []
    for materiel in livraison.materiels:
        materiel.total = materiel.prix_unitaire * materiel.quantite
        materiels_avec_total.append(materiel)
    return render(request, 'livraison/form_livraison_superadmin.html', {
        'form': form,
        'livraison': livraison,
        'materiels': materiels_avec_total,
        'titre': f'Modifier la livraison {livraison.numero_commande}',
        'is_edit': True
    })


@login_required
def supprimer_livraison_superadmin(request, pk):
    livraison = get_object_or_404(Livraison, pk=pk)
    if request.method == 'POST':
        numero = livraison.numero_commande
        livraison.delete()
        messages.success(request, f'Livraison {numero} supprimée avec succès.')
        return redirect('livraison:liste_livraisons_superadmin')
    return render(request, 'livraison/confirmer_suppression_superadmin.html', { 'livraison': livraison })


@login_required
def marquer_livree_superadmin(request, pk):
    livraison = get_object_or_404(Livraison, pk=pk)
    if request.method == 'POST':
        livraison.statut_livraison = 'livree'
        livraison.modifie_par = request.user
        livraison.save()
        messages.success(request, f'Livraison {livraison.numero_commande} marquée comme livrée.')
        return redirect('livraison:detail_livraison_superadmin', pk=pk)
    return render(request, 'livraison/marquer_livree_superadmin.html', { 'livraison': livraison })


@login_required
def valider_pv_reception_superadmin(request, pk):
    livraison = get_object_or_404(Livraison, pk=pk)
    if request.method == 'POST':
        livraison.pv_reception_recu = True
        livraison.modifie_par = request.user
        livraison.save()
        messages.success(request, f'PV de réception validé pour la livraison {livraison.numero_commande}.')
        return redirect('livraison:detail_livraison_superadmin', pk=pk)
    return render(request, 'livraison/valider_pv_reception_superadmin.html', { 'livraison': livraison })


@login_required
def rapports_livraison_superadmin(request):
    user = request.user
    if user.is_superuser:
        livraisons = Livraison.objects.all()
    elif user.groups.filter(name='Gestionnaire Informatique').exists():
        livraisons = Livraison.objects.filter(type_commande='informatique')
    elif user.groups.filter(name='Gestionnaire Bureau').exists():
        livraisons = Livraison.objects.filter(type_commande='bureau')
    else:
        livraisons = Livraison.objects.none()
    context = {
        'total_livraisons': livraisons.count(),
        'livraisons_livrees': livraisons.filter(statut_livraison='livree').count(),
        'livraisons_conformes': livraisons.filter(conforme=True).count(),
        'pv_recus': livraisons.filter(pv_reception_recu=True).count(),
        'livraisons_informatique': livraisons.filter(type_commande='informatique').count(),
        'livraisons_bureau': livraisons.filter(type_commande='bureau').count(),
        'livraisons_par_statut': livraisons.values('statut_livraison').annotate(count=Count('id')).order_by('statut_livraison'),
        'livraisons_par_fournisseur': livraisons.values('commande_informatique__fournisseur__nom').annotate(count=Count('id')).filter(commande_informatique__fournisseur__isnull=False).order_by('-count')[:10],
    }
    return render(request, 'livraison/rapports_superadmin.html', context)


@login_required
def export_livraisons_excel_superadmin(request):
    """Exporter les livraisons (superadmin) en Excel"""
    return export_livraisons_excel(request)


# Vues pour Gestionnaire Informatique
@login_required
def nouvelle_livraison_gestionnaire_info(request):
    """Créer une nouvelle livraison avec formulaire personnalisé (gestionnaire info)"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Info: {user.groups.filter(name='Gestionnaire Informatique').exists()}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions (le superadmin a accès à tout)
    if not (user.groups.filter(name='Gestionnaire Informatique').exists() or user.is_superuser):
        messages.error(request, "Vous n'avez pas les permissions pour créer une livraison.")
        return redirect('livraison:liste_livraison_gestionnaire_info')
    
    # Déterminer le type de commande selon l'utilisateur (toujours informatique pour gestionnaire info)
    type_commande_initial = 'informatique'
    
    if request.method == 'POST':
        form = NouvelleLivraisonForm(request.POST)
        if form.is_valid():
            # Traitement du formulaire
            commande_id = form.cleaned_data['commande_id']
            type_commande = form.cleaned_data['type_commande']
            date_livraison_prevue = form.cleaned_data['date_livraison_prevue']
            date_livraison_effective = form.cleaned_data['date_livraison_effective']
            statut_livraison = form.cleaned_data['statut_livraison']
            conforme = form.cleaned_data['conforme']
            pv_reception_recu = form.cleaned_data['pv_reception_recu']
            notes = form.cleaned_data['notes']
        
            # Utiliser le type de commande sélectionné dans le formulaire
            if type_commande == 'informatique':
                from apps.commande_informatique.models import Commande as CommandeInfo
                try:
                    commande_informatique = CommandeInfo.objects.get(id=commande_id)
                    commande_bureau = None
                except CommandeInfo.DoesNotExist:
                    messages.error(request, "Commande informatique introuvable.")
                    return redirect('livraison:nouvelle_livraison_gestionnaire_info')
            elif type_commande == 'bureau':
                from apps.commande_bureau.models import CommandeBureau
                try:
                    commande_bureau = CommandeBureau.objects.get(id=commande_id)
                    commande_informatique = None
                except CommandeBureau.DoesNotExist:
                    messages.error(request, "Commande bureau introuvable.")
                    return redirect('livraison:nouvelle_livraison_gestionnaire_info')
            else:
                messages.error(request, "Type de commande invalide.")
                return redirect('livraison:nouvelle_livraison_gestionnaire_info')
            
            # Vérifier si une livraison existe déjà pour cette commande
            numero_commande = commande_informatique.numero_commande if commande_informatique else commande_bureau.numero_commande
            
            print(f"DEBUG - Vérification: numero_commande={numero_commande}, type_commande={type_commande}")
            
            livraison_existante = Livraison.objects.filter(numero_commande=numero_commande, type_commande=type_commande).first()
            if livraison_existante:
                print(f"DEBUG - Livraison existante trouvée: {livraison_existante}")
                messages.error(request, f'Une livraison existe déjà pour la commande {numero_commande}.')
                return redirect('livraison:detail_livraison_gestionnaire_info', pk=livraison_existante.pk)
            else:
                print(f"DEBUG - Aucune livraison existante trouvée")
            
            # Créer la livraison
            print(f"DEBUG - Tentative de création: numero_commande={numero_commande}, type_commande={type_commande}")
            try:
                livraison = Livraison.objects.create(
                    numero_commande=numero_commande,
                    type_commande=type_commande,
                    commande_informatique=commande_informatique,
                    commande_bureau=commande_bureau,
                    date_livraison_prevue=date_livraison_prevue,
                    date_livraison_effective=date_livraison_effective or None,
                    statut_livraison=statut_livraison,
                    conforme=conforme,
                    pv_reception_recu=pv_reception_recu,
                    notes=notes,
                    cree_par=user
                )
                
                print(f"DEBUG - Livraison créée avec succès: {livraison}")
                messages.success(request, f'Livraison pour la commande {livraison.numero_commande} créée avec succès.')
                return redirect('livraison:detail_livraison_gestionnaire_info', pk=livraison.pk)
            except Exception as e:
                print(f"DEBUG - Erreur lors de la création: {str(e)}")
                messages.error(request, f'Erreur lors de la création de la livraison : {str(e)}')
                return redirect('livraison:nouvelle_livraison_gestionnaire_info')
        else:
            # Formulaire invalide - afficher les erreurs
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        # GET request - afficher le formulaire
        form = NouvelleLivraisonForm(initial={'type_commande': type_commande_initial})
    
    context = {
        'form': form,
        'type_commande_initial': type_commande_initial
    }
    
    return render(request, 'livraison/form_livraison_gestionnaire_info.html', context)


@login_required
def detail_livraison_gestionnaire_info(request, pk):
    """Détail d'une livraison (gestionnaire info)"""
    livraison = get_object_or_404(Livraison.objects.select_related(
        'commande_informatique__fournisseur',
        'commande_bureau__fournisseur',
        'cree_par',
        'modifie_par'
    ).prefetch_related(
        'commande_informatique__lignes__designation',
        'commande_informatique__lignes__description',
        'commande_bureau__lignes__designation',
        'commande_bureau__lignes__description'
    ), pk=pk)
    
    # Vérifier que c'est une livraison informatique
    if livraison.type_commande != 'informatique':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_info')
    
    materiels_avec_total = []
    for materiel in livraison.materiels:
        materiel.total = materiel.prix_unitaire * materiel.quantite
        materiels_avec_total.append(materiel)
    
    return render(request, 'livraison/detail_livraison_gestionnaire_info.html', {
        'livraison': livraison,
        'materiels': materiels_avec_total,
    })


@login_required
def modifier_livraison_gestionnaire_info(request, pk):
    """Modifier une livraison (gestionnaire info)"""
    livraison = get_object_or_404(Livraison, pk=pk)
    user = request.user
    
    # Vérifier que c'est une livraison informatique
    if livraison.type_commande != 'informatique':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_info')
    
    if request.method == 'POST':
        from .forms import ModifierLivraisonForm
        form = ModifierLivraisonForm(request.POST, instance=livraison)
        if form.is_valid():
            livraison.conforme = form.cleaned_data['conforme']
            livraison.pv_reception_recu = form.cleaned_data['pv_reception_recu']
            livraison.notes = form.cleaned_data['notes']
            livraison.modifie_par = user
            livraison.save()
            messages.success(request, f'Livraison {livraison.numero_commande} modifiée avec succès.')
            return redirect('livraison:detail_livraison_gestionnaire_info', pk=livraison.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        from .forms import ModifierLivraisonForm
        initial_data = {
            'type_commande': livraison.type_commande,
            'commande_id': livraison.commande_informatique.id if livraison.commande_informatique else livraison.commande_bureau.id,
            'date_livraison_prevue': livraison.date_livraison_prevue.strftime('%Y-%m-%d') if livraison.date_livraison_prevue else None,
            'date_livraison_effective': livraison.date_livraison_effective.strftime('%Y-%m-%d') if livraison.date_livraison_effective else None,
            'statut_livraison': livraison.statut_livraison,
            'conforme': livraison.conforme,
            'pv_reception_recu': livraison.pv_reception_recu,
            'notes': livraison.notes,
        }
        form = ModifierLivraisonForm(initial=initial_data)
    
    # Matériels totaux
    materiels_avec_total = []
    for materiel in livraison.materiels:
        materiel.total = materiel.prix_unitaire * materiel.quantite
        materiels_avec_total.append(materiel)
    
    return render(request, 'livraison/form_livraison_gestionnaire_info.html', {
        'form': form,
        'livraison': livraison,
        'materiels': materiels_avec_total,
        'titre': f'Modifier la livraison {livraison.numero_commande}',
        'is_edit': True
    })


@login_required
def supprimer_livraison_gestionnaire_info(request, pk):
    """Supprimer une livraison (gestionnaire info)"""
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison informatique
    if livraison.type_commande != 'informatique':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_info')
    
    if request.method == 'POST':
        numero = livraison.numero_commande
        livraison.delete()
        messages.success(request, f'Livraison {numero} supprimée avec succès.')
        return redirect('livraison:liste_livraisons_gestionnaire_info')
    
    return render(request, 'livraison/confirmer_suppression_gestionnaire_info.html', { 'livraison': livraison })


@login_required
def marquer_livree_gestionnaire_info(request, pk):
    """Marquer une livraison comme livrée (gestionnaire info)"""
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison informatique
    if livraison.type_commande != 'informatique':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_info')
    
    if request.method == 'POST':
        livraison.statut_livraison = 'livree'
        livraison.modifie_par = request.user
        livraison.save()
        messages.success(request, f'Livraison {livraison.numero_commande} marquée comme livrée.')
        return redirect('livraison:detail_livraison_gestionnaire_info', pk=pk)
    
    return render(request, 'livraison/marquer_livree_gestionnaire_info.html', { 'livraison': livraison })


@login_required
def valider_pv_reception_gestionnaire_info(request, pk):
    """Valider le PV de réception (gestionnaire info)"""
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison informatique
    if livraison.type_commande != 'informatique':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_info')
    
    if request.method == 'POST':
        livraison.pv_reception_recu = True
        livraison.modifie_par = request.user
        livraison.save()
        messages.success(request, f'PV de réception validé pour la livraison {livraison.numero_commande}.')
        return redirect('livraison:detail_livraison_gestionnaire_info', pk=pk)
    
    return render(request, 'livraison/valider_pv_reception_gestionnaire_info.html', { 'livraison': livraison })


@login_required
def rapports_livraison_gestionnaire_info(request):
    """Rapports de livraison (gestionnaire info)"""
    user = request.user
    if not user.groups.filter(name='Gestionnaire Informatique').exists():
        messages.error(request, "Vous n'avez pas les permissions pour accéder aux rapports.")
        return redirect('livraison:liste_livraisons_gestionnaire_info')
    
    livraisons = Livraison.objects.filter(type_commande='informatique')
    
    context = {
        'total_livraisons': livraisons.count(),
        'livraisons_livrees': livraisons.filter(statut_livraison='livree').count(),
        'livraisons_conformes': livraisons.filter(conforme=True).count(),
        'pv_recus': livraisons.filter(pv_reception_recu=True).count(),
        'livraisons_informatique': livraisons.filter(type_commande='informatique').count(),
        'livraisons_bureau': 0,  # Toujours 0 pour gestionnaire info
        'livraisons_par_statut': livraisons.values('statut_livraison').annotate(count=Count('id')).order_by('statut_livraison'),
        'livraisons_par_fournisseur': livraisons.values('commande_informatique__fournisseur__nom').annotate(count=Count('id')).filter(commande_informatique__fournisseur__isnull=False).order_by('-count')[:10],
    }
    
    return render(request, 'livraison/rapports_gestionnaire_info.html', context)


@login_required
def export_livraisons_excel_gestionnaire_info(request):
    """Exporter les livraisons informatiques (gestionnaire info) en Excel"""
    user = request.user
    if not user.groups.filter(name='Gestionnaire Informatique').exists():
        messages.error(request, "Vous n'avez pas les permissions pour exporter les données.")
        return redirect('livraison:liste_livraisons_gestionnaire_info')
    
    # Filtrer seulement les livraisons informatiques
    livraisons = Livraison.objects.filter(type_commande='informatique')
    
    # Créer le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="livraisons_informatiques_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Livraisons Informatiques"
    
    # En-têtes
    headers = ['N° Commande', 'Type', 'Fournisseur', 'Date prévue', 'Date effective', 'Statut', 'Conforme', 'PV reçu', 'Notes']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Données
    for row, livraison in enumerate(livraisons, 2):
        ws.cell(row=row, column=1, value=livraison.numero_commande)
        ws.cell(row=row, column=2, value=livraison.get_type_commande_display())
        ws.cell(row=row, column=3, value=livraison.fournisseur.nom if livraison.fournisseur else '-')
        ws.cell(row=row, column=4, value=livraison.date_livraison_prevue.strftime('%d/%m/%Y') if livraison.date_livraison_prevue else '-')
        ws.cell(row=row, column=5, value=livraison.date_livraison_effective.strftime('%d/%m/%Y') if livraison.date_livraison_effective else '-')
        ws.cell(row=row, column=6, value=livraison.get_statut_livraison_display())
        ws.cell(row=row, column=7, value='Oui' if livraison.conforme else 'Non')
        ws.cell(row=row, column=8, value='Oui' if livraison.pv_reception_recu else 'Non')
        ws.cell(row=row, column=9, value=livraison.notes or '-')
    
    wb.save(response)
    return response


# ============================================================================
# VUES GESTIONNAIRE BUREAU
# ============================================================================

@login_required
def liste_livraisons_gestionnaire_bureau(request):
    """Liste des livraisons pour le gestionnaire bureau"""
    user = request.user
    if not (user.groups.filter(name='Gestionnaire Bureau').exists() or user.is_superuser):
        messages.error(request, "Vous n'avez pas les permissions pour accéder à cette page.")
        return redirect('livraison:liste_livraisons')
    
    # Filtrer seulement les livraisons bureau
    livraisons = Livraison.objects.filter(type_commande='bureau').select_related(
        'commande_bureau__fournisseur',
        'cree_par'
    ).prefetch_related(
        'commande_bureau__lignes__designation',
        'commande_bureau__lignes__description'
    ).order_by('-date_creation')
    
    # Filtres
    form = RechercheLivraisonBureauForm(request.GET)
    if form.is_valid():
        search = form.cleaned_data.get('search')
        statut_livraison = form.cleaned_data.get('statut_livraison')
        conforme = form.cleaned_data.get('conforme')
        pv_reception = form.cleaned_data.get('pv_reception')
        
        if search:
            livraisons = livraisons.filter(
                Q(numero_commande__icontains=search) |
                Q(commande_bureau__fournisseur__nom__icontains=search) |
                Q(notes__icontains=search)
            )
        
        if statut_livraison:
            livraisons = livraisons.filter(statut_livraison=statut_livraison)
        
        if conforme and conforme in ['True', 'False']:
            livraisons = livraisons.filter(conforme=conforme == 'True')
        
        if pv_reception and pv_reception in ['True', 'False']:
            livraisons = livraisons.filter(pv_reception_recu=pv_reception == 'True')
    
    # Pagination
    paginator = Paginator(livraisons, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistiques
    total_livraisons = livraisons.count()
    livraisons_livrees = livraisons.filter(statut_livraison='livree').count()
    livraisons_conformes = livraisons.filter(conforme=True).count()
    pv_recus = livraisons.filter(pv_reception_recu=True).count()
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_livraisons': total_livraisons,
        'livraisons_livrees': livraisons_livrees,
        'livraisons_conformes': livraisons_conformes,
        'pv_recus': pv_recus,
        'dashboard_url': 'users:gestionnaire_bureau_dashboard',
    }
    
    return render(request, 'livraison/liste_livraisons_gestionnaire_bureau.html', context)


@login_required
def nouvelle_livraison_gestionnaire_bureau(request):
    """Nouvelle livraison pour le gestionnaire bureau"""
    user = request.user
    if not (user.groups.filter(name='Gestionnaire Bureau').exists() or user.is_superuser):
        messages.error(request, "Vous n'avez pas les permissions pour créer une livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    if request.method == 'POST':
        form = NouvelleLivraisonForm(request.POST)
        if form.is_valid():
            livraison = form.save(commit=False)
            livraison.type_commande = 'bureau'  # Forcer le type bureau
            livraison.cree_par = user
            livraison.save()
            messages.success(request, f'Livraison {livraison.numero_commande} créée avec succès.')
            return redirect('livraison:detail_livraison_gestionnaire_bureau', pk=livraison.pk)
    else:
        form = NouvelleLivraisonForm()
        form.fields['type_commande'].initial = 'bureau'
        form.fields['type_commande'].widget.attrs['readonly'] = True
    
    context = {
        'form': form,
        'dashboard_url': 'users:gestionnaire_bureau_dashboard',
    }
    
    return render(request, 'livraison/form_livraison_gestionnaire_bureau.html', context)


@login_required
def detail_livraison_gestionnaire_bureau(request, pk):
    livraison = get_object_or_404(Livraison.objects.select_related(
        'commande_informatique__fournisseur',
        'commande_bureau__fournisseur',
        'cree_par',
        'modifie_par'
    ).prefetch_related(
        'commande_informatique__lignes__designation',
        'commande_informatique__lignes__description',
        'commande_bureau__lignes__designation',
        'commande_bureau__lignes__description'
    ), pk=pk)
    
    # Vérifier que c'est une livraison bureau
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    materiels_avec_total = []
    for materiel in livraison.materiels:
        materiel.total = materiel.prix_unitaire * materiel.quantite
        materiels_avec_total.append(materiel)
    
    return render(request, 'livraison/detail_livraison_gestionnaire_bureau.html', {
        'livraison': livraison,
        'materiels': materiels_avec_total,
    })


@login_required
def modifier_livraison_gestionnaire_bureau(request, pk):
    livraison = get_object_or_404(Livraison, pk=pk)
    user = request.user
    
    # Vérifier que c'est une livraison bureau
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    if request.method == 'POST':
        from .forms import ModifierLivraisonForm
        form = ModifierLivraisonForm(request.POST)
        if form.is_valid():
            livraison.date_livraison_prevue = form.cleaned_data['date_livraison_prevue']
            livraison.date_livraison_effective = form.cleaned_data['date_livraison_effective'] or None
            livraison.statut_livraison = form.cleaned_data['statut_livraison']
            livraison.conforme = form.cleaned_data['conforme']
            livraison.pv_reception_recu = form.cleaned_data['pv_reception_recu']
            livraison.notes = form.cleaned_data['notes']
            livraison.modifie_par = user
            livraison.save()
            messages.success(request, f'Livraison {livraison.numero_commande} modifiée avec succès.')
            return redirect('livraison:detail_livraison_gestionnaire_bureau', pk=livraison.pk)
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        from .forms import ModifierLivraisonForm
        initial_data = {
            'type_commande': livraison.type_commande,
            'commande_id': livraison.commande_informatique.id if livraison.commande_informatique else livraison.commande_bureau.id,
            'date_livraison_prevue': livraison.date_livraison_prevue.strftime('%Y-%m-%d') if livraison.date_livraison_prevue else None,
            'date_livraison_effective': livraison.date_livraison_effective.strftime('%Y-%m-%d') if livraison.date_livraison_effective else None,
            'statut_livraison': livraison.statut_livraison,
            'conforme': livraison.conforme,
            'pv_reception_recu': livraison.pv_reception_recu,
            'notes': livraison.notes,
        }
        form = ModifierLivraisonForm(initial=initial_data)
    
    # Matériels totaux
    materiels_avec_total = []
    for materiel in livraison.materiels:
        materiel.total = materiel.prix_unitaire * materiel.quantite
        materiels_avec_total.append(materiel)
    
    return render(request, 'livraison/modifier_livraison_gestionnaire_bureau.html', {
        'form': form,
        'livraison': livraison,
        'materiels': materiels_avec_total,
        'titre': f'Modifier la livraison {livraison.numero_commande}',
        'is_edit': True
    })


@login_required
def marquer_livree_gestionnaire_bureau(request, pk):
    """Marquer une livraison comme livrée (gestionnaire bureau)"""
    user = request.user
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour effectuer cette action.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison bureau
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    if request.method == 'POST':
        livraison.statut_livraison = 'livree'
        livraison.modifie_par = user
        livraison.save()
        messages.success(request, f'Livraison {livraison.numero_commande} marquée comme livrée.')
        return redirect('livraison:detail_livraison_gestionnaire_bureau', pk=pk)
    
    return render(request, 'livraison/marquer_livree_gestionnaire_bureau.html', { 'livraison': livraison })


@login_required
def valider_pv_reception_gestionnaire_bureau(request, pk):
    """Valider le PV de réception (gestionnaire bureau)"""
    user = request.user
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour effectuer cette action.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison bureau
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    if request.method == 'POST':
        livraison.pv_reception_recu = True
        livraison.modifie_par = user
        livraison.save()
        messages.success(request, f'PV de réception validé pour la livraison {livraison.numero_commande}.')
        return redirect('livraison:detail_livraison_gestionnaire_bureau', pk=pk)
    
    return render(request, 'livraison/valider_pv_reception_gestionnaire_bureau.html', { 'livraison': livraison })


@login_required
def supprimer_livraison_gestionnaire_bureau(request, pk):
    """Supprimer une livraison (gestionnaire bureau)"""
    user = request.user
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour supprimer cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison bureau
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    if request.method == 'POST':
        numero_commande = livraison.numero_commande
        livraison.delete()
        messages.success(request, f'Livraison {numero_commande} supprimée avec succès.')
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    return render(request, 'livraison/confirmer_suppression_gestionnaire_bureau.html', { 'livraison': livraison })


@login_required
def rapports_livraison_gestionnaire_bureau(request):
    """Rapports de livraison (gestionnaire bureau)"""
    user = request.user
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour accéder aux rapports.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    livraisons = Livraison.objects.filter(type_commande='bureau')
    
    context = {
        'total_livraisons': livraisons.count(),
        'livraisons_livrees': livraisons.filter(statut_livraison='livree').count(),
        'livraisons_conformes': livraisons.filter(conforme=True).count(),
        'pv_recus': livraisons.filter(pv_reception_recu=True).count(),
        'livraisons_informatique': 0,  # Toujours 0 pour gestionnaire bureau
        'livraisons_bureau': livraisons.filter(type_commande='bureau').count(),
        'livraisons_par_statut': livraisons.values('statut_livraison').annotate(count=Count('id')).order_by('statut_livraison'),
        'livraisons_par_fournisseur': livraisons.values('commande_bureau__fournisseur__nom').annotate(count=Count('id')).filter(commande_bureau__fournisseur__isnull=False).order_by('-count')[:10],
    }
    
    return render(request, 'livraison/rapports_gestionnaire_bureau.html', context)


@login_required
def export_livraisons_excel_gestionnaire_bureau(request):
    """Exporter les livraisons bureau (gestionnaire bureau) en Excel"""
    user = request.user
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour exporter les données.")
        return redirect('livraison:liste_livraisons_gestionnaire_bureau')
    
    # Filtrer seulement les livraisons bureau
    livraisons = Livraison.objects.filter(type_commande='bureau')
    
    # Créer le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="livraisons_bureau_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Livraisons Bureau"
    
    # En-têtes
    headers = ['N° Commande', 'Type', 'Fournisseur', 'Date prévue', 'Date effective', 'Statut', 'Conforme', 'PV reçu', 'Notes']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Données
    for row, livraison in enumerate(livraisons, 2):
        ws.cell(row=row, column=1, value=livraison.numero_commande)
        ws.cell(row=row, column=2, value=livraison.get_type_commande_display())
        ws.cell(row=row, column=3, value=livraison.fournisseur.nom if livraison.fournisseur else '-')
        ws.cell(row=row, column=4, value=livraison.date_livraison_prevue.strftime('%d/%m/%Y') if livraison.date_livraison_prevue else '-')
        ws.cell(row=row, column=5, value=livraison.date_livraison_effective.strftime('%d/%m/%Y') if livraison.date_livraison_effective else '-')
        ws.cell(row=row, column=6, value=livraison.get_statut_livraison_display())
        ws.cell(row=row, column=7, value='Oui' if livraison.conforme else 'Non')
        ws.cell(row=row, column=8, value='Oui' if livraison.pv_reception_recu else 'Non')
        ws.cell(row=row, column=9, value=livraison.notes or '-')
    
    wb.save(response)
    return response

# ============================================================================
# VUES LIVRAISONS BUREAU GESTIONNAIRE BUREAU (même code que gestionnaire info, mais pour commandes bureau)
# ============================================================================





@login_required
def detail_livraison_informatique_gestionnaire_bureau(request, pk):
    """Détail d'une livraison bureau (gestionnaire bureau) - même code que gestionnaire info"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour accéder à cette livraison.")
        return redirect('users:gestionnaire_bureau_dashboard')
    
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison bureau (gestionnaire bureau gère seulement les commandes bureau)
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_informatique_gestionnaire_bureau')
    
    context = {
        'livraison': livraison,
        'dashboard_url': 'users:gestionnaire_bureau_dashboard',
    }
    
    return render(request, 'livraison/detail_livraison_informatique_gestionnaire_bureau.html', context)

@login_required
def modifier_livraison_informatique_gestionnaire_bureau(request, pk):
    """Modifier une livraison bureau (gestionnaire bureau) - même code que gestionnaire info"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour modifier cette livraison.")
        return redirect('users:gestionnaire_bureau_dashboard')
    
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison bureau (gestionnaire bureau gère seulement les commandes bureau)
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_informatique_gestionnaire_bureau')
    
    if request.method == 'POST':
        form = LivraisonForm(request.POST, instance=livraison)
        if form.is_valid():
            livraison = form.save(commit=False)
            livraison.modifie_par = user
            livraison.save()
            messages.success(request, f'Livraison {livraison.numero_commande} modifiée avec succès.')
            return redirect('livraison:detail_livraison_informatique_gestionnaire_bureau', pk=pk)
    else:
        form = LivraisonForm(instance=livraison)
    
    context = {
        'form': form,
        'livraison': livraison,
        'dashboard_url': 'users:gestionnaire_bureau_dashboard',
    }
    
    return render(request, 'livraison/modifier_livraison_informatique_gestionnaire_bureau.html', context)

@login_required
def marquer_livree_informatique_gestionnaire_bureau(request, pk):
    """Marquer une livraison bureau comme livrée (gestionnaire bureau) - même code que gestionnaire info"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour modifier cette livraison.")
        return redirect('users:gestionnaire_bureau_dashboard')
    
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison bureau (gestionnaire bureau gère seulement les commandes bureau)
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_informatique_gestionnaire_bureau')
    
    if request.method == 'POST':
        livraison.statut_livraison = 'livree'
        livraison.modifie_par = user
        livraison.save()
        messages.success(request, f'Livraison {livraison.numero_commande} marquée comme livrée.')
        return redirect('livraison:detail_livraison_informatique_gestionnaire_bureau', pk=pk)
    
    return render(request, 'livraison/marquer_livree_informatique_gestionnaire_bureau.html', { 'livraison': livraison })

@login_required
def valider_pv_reception_informatique_gestionnaire_bureau(request, pk):
    """Valider le PV de réception d'une livraison bureau (gestionnaire bureau) - même code que gestionnaire info"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour valider ce PV.")
        return redirect('users:gestionnaire_bureau_dashboard')
    
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison bureau (gestionnaire bureau gère seulement les commandes bureau)
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_informatique_gestionnaire_bureau')
    
    if request.method == 'POST':
        livraison.pv_reception_recu = True
        livraison.modifie_par = user
        livraison.save()
        messages.success(request, f'PV de réception validé pour la livraison {livraison.numero_commande}.')
        return redirect('livraison:detail_livraison_informatique_gestionnaire_bureau', pk=pk)
    
    return render(request, 'livraison/valider_pv_reception_informatique_gestionnaire_bureau.html', { 'livraison': livraison })

@login_required
def supprimer_livraison_informatique_gestionnaire_bureau(request, pk):
    """Supprimer une livraison bureau (gestionnaire bureau) - même code que gestionnaire info"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour supprimer cette livraison.")
        return redirect('users:gestionnaire_bureau_dashboard')
    
    livraison = get_object_or_404(Livraison, pk=pk)
    
    # Vérifier que c'est une livraison bureau (gestionnaire bureau gère seulement les commandes bureau)
    if livraison.type_commande != 'bureau':
        messages.error(request, "Vous n'avez pas accès à cette livraison.")
        return redirect('livraison:liste_livraisons_informatique_gestionnaire_bureau')
    
    if request.method == 'POST':
        numero_commande = livraison.numero_commande
        livraison.delete()
        messages.success(request, f'Livraison {numero_commande} supprimée avec succès.')
        return redirect('livraison:liste_livraisons_informatique_gestionnaire_bureau')
    
    return render(request, 'livraison/confirmer_suppression_informatique_gestionnaire_bureau.html', { 'livraison': livraison })

@login_required
def rapports_livraison_informatique_gestionnaire_bureau(request):
    """Rapports de livraison bureau (gestionnaire bureau) - même code que gestionnaire info"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour accéder aux rapports.")
        return redirect('users:gestionnaire_bureau_dashboard')
    
    # Filtrer seulement les livraisons bureau (gestionnaire bureau gère seulement les commandes bureau)
    livraisons = Livraison.objects.filter(type_commande='bureau')
    
    context = {
        'total_livraisons': livraisons.count(),
        'livraisons_livrees': livraisons.filter(statut_livraison='livree').count(),
        'livraisons_conformes': livraisons.filter(conforme=True).count(),
        'pv_recus': livraisons.filter(pv_reception_recu=True).count(),
        'livraisons_informatique': 0,  # Toujours 0 pour gestionnaire bureau
        'livraisons_bureau': livraisons.filter(type_commande='bureau').count(),
        'livraisons_par_statut': livraisons.values('statut_livraison').annotate(count=Count('id')).order_by('statut_livraison'),
        'livraisons_par_fournisseur': livraisons.values('commande_bureau__fournisseur__nom').annotate(count=Count('id')).filter(commande_bureau__fournisseur__isnull=False).order_by('-count')[:10],
    }
    
    return render(request, 'livraison/rapports_informatique_gestionnaire_bureau.html', context)

@login_required
def export_livraisons_excel_informatique_gestionnaire_bureau(request):
    """Exporter les livraisons bureau (gestionnaire bureau) en Excel - même code que gestionnaire info"""
    user = request.user
    
    # Debug: afficher les informations de l'utilisateur
    print(f"DEBUG - Utilisateur: {user.username}")
    print(f"DEBUG - Superuser: {user.is_superuser}")
    print(f"DEBUG - Groupes: {list(user.groups.all())}")
    print(f"DEBUG - Gestionnaire Bureau: {user.groups.filter(name='Gestionnaire Bureau').exists()}")
    
    # Vérifier les permissions
    if not user.groups.filter(name='Gestionnaire Bureau').exists():
        messages.error(request, "Vous n'avez pas les permissions pour exporter les données.")
        return redirect('users:gestionnaire_bureau_dashboard')
    
    # Filtrer seulement les livraisons bureau (gestionnaire bureau gère seulement les commandes bureau)
    livraisons = Livraison.objects.filter(type_commande='bureau')
    
    # Créer le fichier Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="livraisons_bureau_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Livraisons Bureau"
    
    # En-têtes
    headers = ['N° Commande', 'Type', 'Fournisseur', 'Date prévue', 'Date effective', 'Statut', 'Conforme', 'PV reçu', 'Notes']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Données
    for row, livraison in enumerate(livraisons, 2):
        ws.cell(row=row, column=1, value=livraison.numero_commande)
        ws.cell(row=row, column=2, value=livraison.get_type_commande_display())
        ws.cell(row=row, column=3, value=livraison.fournisseur.nom if livraison.fournisseur else '-')
        ws.cell(row=row, column=4, value=livraison.date_livraison_prevue.strftime('%d/%m/%Y') if livraison.date_livraison_prevue else '-')
        ws.cell(row=row, column=5, value=livraison.date_livraison_effective.strftime('%d/%m/%Y') if livraison.date_livraison_effective else '-')
        ws.cell(row=row, column=6, value=livraison.get_statut_livraison_display())
        ws.cell(row=row, column=7, value='Oui' if livraison.conforme else 'Non')
        ws.cell(row=row, column=8, value='Oui' if livraison.pv_reception_recu else 'Non')
        ws.cell(row=row, column=9, value=livraison.notes or '-')
    
    wb.save(response)
    return response
