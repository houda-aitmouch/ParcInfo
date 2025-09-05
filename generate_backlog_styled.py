#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Générateur de Backlog Produit Stylé - ParcInfo
Crée un backlog produit structuré avec style professionnel
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.worksheet.worksheet import Worksheet

def apply_styles_to_worksheet(ws, title="Backlog Produit"):
    """Applique des styles professionnels à la feuille de calcul"""
    
    # Styles de base
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titre principal
    ws['A1'] = f"📋 {title} - ParcInfo"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
    ws.merge_cells('A1:L1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Sous-titre avec date
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws.merge_cells('A2:L2')
    ws['A2'].alignment = Alignment(horizontal='center')
    
    # En-têtes de colonnes
    headers = ['Prio', 'EPIC', 'Feature', 'Persona', 'Macro US', 'RG', 'US', 'Dépendance', 'Statut', 'Commentaire', 'Feature Link', 'US Link']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Ajuster la largeur des colonnes
    column_widths = {
        'A': 8,   # Prio
        'B': 20,  # EPIC
        'C': 25,  # Feature
        'D': 15,  # Persona
        'E': 40,  # Macro US
        'F': 20,  # RG
        'G': 25,  # US
        'H': 15,  # Dépendance
        'I': 12,  # Statut
        'J': 30,  # Commentaire
        'K': 20,  # Feature Link
        'L': 20   # US Link
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Hauteur de ligne pour les en-têtes
    ws.row_dimensions[4].height = 30
    
    return ws

def apply_priority_colors(ws, start_row=5):
    """Applique des couleurs selon la priorité"""
    
    # Couleurs par priorité
    priority_colors = {
        'P1': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),  # Rouge clair
        'P2': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),  # Jaune clair
        'P3': PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')   # Bleu clair
    }
    
    # Couleurs par statut
    status_colors = {
        'Terminé': PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid'),  # Vert
        'En cours': PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid'),  # Jaune
        'À faire': PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')    # Rouge
    }
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, ws.max_row + 1):
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Couleur de priorité pour la première colonne
            if col == 1 and cell.value in priority_colors:
                cell.fill = priority_colors[cell.value]
                cell.font = Font(bold=True)
            
            # Couleur de statut pour la colonne statut
            if col == 9 and cell.value in status_colors:
                cell.fill = status_colors[cell.value]
                cell.font = Font(bold=True)

def create_personas_sheet(wb):
    """Crée la feuille des personas avec style"""
    ws = wb.create_sheet("Personas")
    
    # Titre
    ws['A1'] = "👥 Personas - ParcInfo"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # En-têtes
    headers = ['Persona', 'Rôle', 'Responsabilités', 'Accès', 'Permissions']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Données des personas
    personas_data = [
        ['Employé', 'Utilisateur final du système', 'Créer des demandes, signer des décharges, consulter ses équipements', 'Dashboard employé, demandes personnelles, notifications', 'Lecture sur ses données, écriture sur ses demandes'],
        ['Gestionnaire Info', 'Gestionnaire du parc informatique', 'Approuver demandes IT, gérer matériel informatique, suivre commandes IT', 'Dashboard gestionnaire info, matériel informatique, commandes IT', 'Lecture/écriture sur matériel IT, approbation demandes IT'],
        ['Gestionnaire Bureau', 'Gestionnaire du mobilier et fournitures de bureau', 'Approuver demandes bureau, gérer mobilier, suivre commandes bureau', 'Dashboard gestionnaire bureau, matériel bureau, commandes bureau', 'Lecture/écriture sur matériel bureau, approbation demandes bureau'],
        ['Super Admin', 'Administrateur système', 'Gestion complète du système, utilisateurs, fournisseurs', 'Toutes les fonctionnalités, configuration système', 'Accès complet à toutes les fonctionnalités']
    ]
    
    for row_idx, persona in enumerate(personas_data, 4):
        for col_idx, value in enumerate(persona, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Largeur des colonnes
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 35

def create_epics_sheet(wb):
    """Crée la feuille des EPICs avec style"""
    ws = wb.create_sheet("EPICs")
    
    # Titre
    ws['A1'] = "📋 EPICs - ParcInfo"
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='366092')
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # En-têtes
    headers = ['EPIC', 'Description', 'Responsable']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Données des EPICs
    epics_data = [
        ['Authentification', 'Gestion des connexions et sessions utilisateurs', 'Super Admin'],
        ['Gestion Utilisateurs', 'Gestion des profils et rôles utilisateurs', 'Super Admin'],
        ['Demandes Équipement', 'Workflow de demandes d\'équipement', 'Gestionnaire Info/Bureau'],
        ['Matériel Informatique', 'Gestion du parc informatique', 'Gestionnaire Info'],
        ['Matériel Bureau', 'Gestion du mobilier et fournitures', 'Gestionnaire Bureau'],
        ['Commandes Informatiques', 'Gestion des commandes informatiques', 'Gestionnaire Info'],
        ['Commandes Bureau', 'Gestion des commandes bureau', 'Gestionnaire Bureau'],
        ['Fournisseurs', 'Gestion des fournisseurs et catalogues', 'Super Admin'],
        ['Livraisons', 'Suivi des livraisons et réceptions', 'Gestionnaire Info'],
        ['Notifications', 'Système de notifications et alertes', 'Super Admin'],
        ['Dashboard', 'Tableaux de bord personnalisés', 'Super Admin'],
        ['Chatbot', 'Assistant IA et base de connaissances', 'Super Admin'],
        ['Rapports', 'Génération de rapports et exports', 'Gestionnaire Info'],
        ['Recherche', 'Recherche globale et filtres', 'Super Admin'],
        ['Garanties', 'Suivi des garanties et alertes', 'Gestionnaire Info'],
        ['Archives', 'Archivage électronique sécurisé', 'Super Admin'],
        ['API', 'API REST pour intégrations', 'Super Admin'],
        ['Sécurité', 'Audit et sécurité du système', 'Super Admin'],
        ['Mobile', 'Application mobile et PWA', 'Super Admin'],
        ['Intégration', 'Intégrations avec systèmes externes', 'Super Admin'],
        ['BI', 'Business Intelligence et analytics', 'Super Admin']
    ]
    
    for row_idx, epic in enumerate(epics_data, 4):
        for col_idx, value in enumerate(epic, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Largeur des colonnes
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20

def generate_styled_backlog():
    """Génère le backlog produit avec style professionnel"""
    
    # Données du backlog (version simplifiée pour l'exemple)
    backlog_data = [
        {'Prio': 'P1', 'EPIC': 'Authentification', 'Feature': 'Interface de connexion', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux me connecter au système', 'RG': 'RG-AUTH-001: Champs identifiant et mot de passe obligatoires', 'US': 'US-AUTH-001: Saisir mon identifiant et mot de passe', 'Dépendance': '-', 'Statut': 'Terminé', 'Commentaire': 'Interface responsive et sécurisée', 'Feature Link': '/users/login/', 'US Link': '/users/login/'},
        {'Prio': 'P1', 'EPIC': 'Demandes Équipement', 'Feature': 'Création de demande', 'Persona': 'Employé', 'Macro US': 'En tant qu\'employé, je veux créer une demande d\'équipement', 'RG': 'RG-DEM-001: Champs obligatoires selon le type', 'US': 'US-DEM-001: Remplir le formulaire de demande', 'Dépendance': 'Authentification', 'Statut': 'Terminé', 'Commentaire': 'Formulaire dynamique selon catégorie', 'Feature Link': '/demande_equipement/nouvelle/', 'US Link': '/demande_equipement/nouvelle/'},
        {'Prio': 'P1', 'EPIC': 'Matériel Informatique', 'Feature': 'Gestion du parc', 'Persona': 'Gestionnaire Info', 'Macro US': 'En tant que gestionnaire, je veux gérer le matériel informatique', 'RG': 'RG-MAT-001: Inventaire complet avec codes uniques', 'US': 'US-MAT-001: Ajouter/modifier/supprimer du matériel', 'Dépendance': '-', 'Statut': 'Terminé', 'Commentaire': 'Codes d\'inventaire automatiques', 'Feature Link': '/materiel_informatique/', 'US Link': '/materiel_informatique/'},
        {'Prio': 'P2', 'EPIC': 'Chatbot', 'Feature': 'Assistant IA', 'Persona': 'Employé', 'Macro US': 'En tant qu\'utilisateur, je veux poser des questions', 'RG': 'RG-CHAT-001: Réponses basées sur la base de connaissances', 'US': 'US-CHAT-001: Interroger l\'assistant IA', 'Dépendance': '-', 'Statut': 'Terminé', 'Commentaire': 'RAG avec base de connaissances', 'Feature Link': '/chatbot/', 'US Link': '/chatbot/'},
        {'Prio': 'P3', 'EPIC': 'API', 'Feature': 'API REST', 'Persona': 'Super Admin', 'Macro US': 'En tant qu\'admin, je veux une API', 'RG': 'RG-API-001: Documentation OpenAPI', 'US': 'US-API-001: Exposer les données via API', 'Dépendance': '-', 'Statut': 'À faire', 'Commentaire': 'API RESTful complète', 'Feature Link': '/api/', 'US Link': '/api/'}
    ]
    
    df = pd.DataFrame(backlog_data)
    
    # Création du workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Backlog Produit"
    
    # Application des styles
    apply_styles_to_worksheet(ws)
    
    # Ajout des données
    for r in dataframe_to_rows(df, index=False, header=False):
        ws.append(r)
    
    # Application des couleurs
    apply_priority_colors(ws, start_row=5)
    
    # Création des autres feuilles
    create_personas_sheet(wb)
    create_epics_sheet(wb)
    
    # Suppression de la feuille par défaut si elle existe
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Sauvegarde
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Backlog_Produit_ParcInfo_Styled_{timestamp}.xlsx"
    wb.save(filename)
    
    print(f"✅ Backlog produit stylé généré : {filename}")
    print(f"📊 {len(df)} fonctionnalités avec style professionnel")
    print(f"🎨 Couleurs par priorité et statut appliquées")
    
    return filename

if __name__ == "__main__":
    generate_styled_backlog()
